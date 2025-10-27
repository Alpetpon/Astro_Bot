from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.filters import Command, StateFilter
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from bson import ObjectId
import os
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

from config import config
from database import get_db, User, Payment, UserRepository, PaymentRepository
from keyboards import get_admin_keyboard, get_bot_management_keyboard, get_back_to_admin_keyboard
from data import get_all_courses, get_all_consultations
from utils.bot_settings import is_admin

# –ó–∞–≥–ª—É—à–∫–∏ –¥–ª—è —Å—Ç–∞—Ä—ã—Ö –º–æ–¥–µ–ª–µ–π –ë–î (–∫–æ–¥ –Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è, –ø—Ä–æ–µ–∫—Ç —Ä–∞–±–æ—Ç–∞–µ—Ç –Ω–∞ JSON)
Course = None  # type: ignore
Consultation = None  # type: ignore
Lesson = None  # type: ignore

router = Router()


class CourseManagement(StatesGroup):
    """–°–æ—Å—Ç–æ—è–Ω–∏—è –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –∫—É—Ä—Å–∞–º–∏"""
    waiting_for_lesson_title = State()
    waiting_for_lesson_description = State()
    waiting_for_lesson_content = State()
    waiting_for_lesson_video = State()
    waiting_for_lesson_lecture_file = State()
    waiting_for_lesson_text_content = State()
    editing_lesson_title = State()
    editing_lesson_description = State()
    editing_lesson_content = State()
    editing_lesson_video = State()


class ConsultationManagement(StatesGroup):
    """–°–æ—Å—Ç–æ—è–Ω–∏—è –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è–º–∏"""
    editing_name = State()
    editing_price = State()
    editing_duration = State()
    editing_description = State()
    # –°–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–π –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏
    creating_name = State()
    creating_slug = State()
    creating_emoji = State()
    creating_price = State()
    creating_duration = State()
    creating_description = State()


class CourseCreation(StatesGroup):
    """–°–æ—Å—Ç–æ—è–Ω–∏—è –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –Ω–æ–≤–æ–≥–æ –∫—É—Ä—Å–∞"""
    waiting_for_name = State()
    waiting_for_slug = State()
    waiting_for_description = State()
    waiting_for_duration = State()


class PaymentLinkCreation(StatesGroup):
    """–°–æ—Å—Ç–æ—è–Ω–∏—è –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –ø–ª–∞—Ç–µ–∂–Ω–æ–π —Å—Å—ã–ª–∫–∏"""
    waiting_for_user_id = State()
    waiting_for_product_type = State()
    waiting_for_product_selection = State()
    waiting_for_amount = State()


class BroadcastStates(StatesGroup):
    """–°–æ—Å—Ç–æ—è–Ω–∏—è –¥–ª—è —Ä–∞—Å—Å—ã–ª–∫–∏"""
    waiting_for_text = State()
    waiting_for_photo = State()
    waiting_for_video = State()
    waiting_for_caption = State()


@router.message(Command("admin"))
async def cmd_admin(message: Message):
    """–í—Ö–æ–¥ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å –ø–æ –∫–æ–º–∞–Ω–¥–µ /admin"""
    if not is_admin(message.from_user.id):
        await message.answer("‚ùå –£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏")
        return
    
    await message.answer(
        "üîê <b>–ê–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å</b>\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=get_admin_keyboard()
    )


@router.callback_query(F.data == "admin_panel")
async def show_admin_panel(callback: CallbackQuery):
    """–ü–æ–∫–∞–∑–∞—Ç—å –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await callback.message.edit_text(
        "üîê <b>–ê–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å</b>\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=get_admin_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "admin_bot_management")
async def show_bot_management(callback: CallbackQuery):
    """–ü–æ–∫–∞–∑–∞—Ç—å –ø–æ–¥–º–µ–Ω—é —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –±–æ—Ç–æ–º"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await callback.message.edit_text(
        "ü§ñ <b>–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –±–æ—Ç–æ–º</b>\n\n"
        "–í—ã–±–µ—Ä–∏—Ç–µ —Ä–∞–∑–¥–µ–ª –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è:",
        reply_markup=get_bot_management_keyboard()
    )
    await callback.answer()


@router.callback_query(F.data == "admin_stats")
async def show_stats(callback: CallbackQuery):
    """–ü–æ–∫–∞–∑–∞—Ç—å —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    db = await get_db()
    user_repo = UserRepository(db)
    payment_repo = PaymentRepository(db)
    
    # –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
    total_users = await user_repo.count()
    total_purchases = await payment_repo.count_by_status("succeeded")
    total_revenue = await payment_repo.sum_by_status("succeeded")
    
    # –ê–∫—Ç–∏–≤–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ (–∑–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–µ 7 –¥–Ω–µ–π)
    week_ago = datetime.utcnow() - timedelta(days=7)
    active_users = await user_repo.count_active_since(week_ago)
    
    # –ù–æ–≤—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ –∑–∞ –Ω–µ–¥–µ–ª—é
    new_users = await user_repo.count_created_since(week_ago)
    
    # –ü–æ–∫—É–ø–∫–∏ –∑–∞ –Ω–µ–¥–µ–ª—é
    week_purchases = await payment_repo.count_since(week_ago, "succeeded")
    week_revenue = await payment_repo.sum_since(week_ago, "succeeded")
    
    # –ö—É—Ä—Å—ã, –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ –∏ –≥–∞–π–¥—ã (–∏–∑ JSON)
    total_courses = len(get_all_courses())
    total_consultations = len(get_all_consultations())
    
    from data import get_all_guides
    total_guides = len(get_all_guides())
    
    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ–¥–ø–∏—Å–æ–∫ (–µ—Å–ª–∏ —Å–µ—Ä–≤–∏—Å –¥–æ—Å—Ç—É–ø–µ–Ω)
    subscription_stats = None
    try:
        from handlers.subscription_handlers import subscription_service
        if subscription_service:
            subscription_stats = await subscription_service.get_subscription_stats()
    except Exception as e:
        logger.warning(f"Could not get subscription stats: {e}")
    
    stats_text = f"""üìä <b>–°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞</b>

üë• <b>–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏:</b>
‚Ä¢ –í—Å–µ–≥–æ: {total_users}
‚Ä¢ –ê–∫—Ç–∏–≤–Ω—ã—Ö –∑–∞ –Ω–µ–¥–µ–ª—é: {active_users}
‚Ä¢ –ù–æ–≤—ã—Ö –∑–∞ –Ω–µ–¥–µ–ª—é: {new_users}

üí∞ <b>–§–∏–Ω–∞–Ω—Å—ã:</b>
‚Ä¢ –í—Å–µ–≥–æ –ø–æ–∫—É–ø–æ–∫: {total_purchases}
‚Ä¢ –û–±—â–∞—è –≤—ã—Ä—É—á–∫–∞: {total_revenue:,.0f} ‚ÇΩ
‚Ä¢ –ó–∞ –Ω–µ–¥–µ–ª—é –ø–æ–∫—É–ø–æ–∫: {week_purchases}
‚Ä¢ –ó–∞ –Ω–µ–¥–µ–ª—é –≤—ã—Ä—É—á–∫–∞: {week_revenue:,.0f} ‚ÇΩ

üìö <b>–ö–æ–Ω—Ç–µ–Ω—Ç:</b>
‚Ä¢ –ö—É—Ä—Å–æ–≤: {total_courses}
‚Ä¢ –ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π: {total_consultations}
‚Ä¢ –ì–∞–π–¥–æ–≤: {total_guides}"""
    
    # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É –ø–æ–¥–ø–∏—Å–æ–∫, –µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–Ω–∞
    if subscription_stats:
        stats_text += f"""

üí´ <b>–ü–æ–¥–ø–∏—Å–∫–∏ –Ω–∞ –∫–∞–Ω–∞–ª:</b>
‚Ä¢ –í—Å–µ–≥–æ –ø–æ–¥–ø–∏—Å–æ–∫: {subscription_stats['total_subscriptions']}
‚Ä¢ –ê–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–¥–ø–∏—Å–æ–∫: {subscription_stats['active_subscriptions']}
‚Ä¢ –í—Å–µ–≥–æ –ø–ª–∞—Ç–µ–∂–µ–π: {subscription_stats['total_payments']}
‚Ä¢ –£—Å–ø–µ—à–Ω—ã—Ö –ø–ª–∞—Ç–µ–∂–µ–π: {subscription_stats['succeeded_payments']}
‚Ä¢ –û–±—â–∞—è —Å—É–º–º–∞: {subscription_stats['total_amount']:,.0f} ‚ÇΩ
"""
    
    # –ö–Ω–æ–ø–∫–∏ —Å –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—å—é —Å–∫–∞—á–∞—Ç—å Excel
    buttons = [
        [InlineKeyboardButton(text="üì• –°–∫–∞—á–∞—Ç—å –¥–µ—Ç–∞–ª—å–Ω—É—é –∞–Ω–∞–ª–∏—Ç–∏–∫—É (Excel)", callback_data="download_analytics")],
        [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å", callback_data="admin_panel")]
    ]
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(
        stats_text,
        reply_markup=keyboard
    )
    
    await callback.answer()


@router.callback_query(F.data == "admin_broadcast")
async def start_broadcast(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞—Ç—å —Ä–∞—Å—Å—ã–ª–∫—É"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await callback.message.edit_text(
        "üì¢ <b>–†–∞—Å—Å—ã–ª–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –¥–ª—è —Ä–∞—Å—Å—ã–ª–∫–∏:\n"
        "‚Ä¢ –¢–µ–∫—Å—Ç\n"
        "‚Ä¢ –§–æ—Ç–æ (—Å –ø–æ–¥–ø–∏—Å—å—é –∏–ª–∏ –±–µ–∑)\n"
        "‚Ä¢ –í–∏–¥–µ–æ (—Å –ø–æ–¥–ø–∏—Å—å—é –∏–ª–∏ –±–µ–∑)\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ /cancel –¥–ª—è –æ—Ç–º–µ–Ω—ã."
    )
    await state.set_state(BroadcastStates.waiting_for_text)
    await callback.answer()


@router.message(Command("broadcast"))
async def broadcast_message(message: Message):
    """–†–∞—Å—Å—ã–ª–∫–∞ —Å–æ–æ–±—â–µ–Ω–∏—è –≤—Å–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è–º"""
    if not is_admin(message.from_user.id):
        await message.answer("‚ùå –£ –≤–∞—Å –Ω–µ—Ç –¥–æ—Å—Ç—É–ø–∞ –∫ —ç—Ç–æ–π –∫–æ–º–∞–Ω–¥–µ")
        return
    
    # –ü–æ–ª—É—á–∞–µ–º —Ç–µ–∫—Å—Ç —Å–æ–æ–±—â–µ–Ω–∏—è
    text = message.text.replace("/broadcast", "").strip()
    
    if not text:
        await message.answer(
            "‚ùå –£–∫–∞–∂–∏—Ç–µ —Ç–µ–∫—Å—Ç —Å–æ–æ–±—â–µ–Ω–∏—è:\n"
            "<code>/broadcast [—Ç–µ–∫—Å—Ç]</code>"
        )
        return
    
    db = await get_db()
    user_repo = UserRepository(db)
    
    # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
    users = await user_repo.get_all()
    
    success_count = 0
    fail_count = 0
    
    status_msg = await message.answer(f"üì§ –ù–∞—á–∏–Ω–∞—é —Ä–∞—Å—Å—ã–ª–∫—É –¥–ª—è {len(users)} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π...")
    
    for user in users:
        try:
            await message.bot.send_message(
                chat_id=user.telegram_id,
                text=text
            )
            success_count += 1
        except Exception:
            fail_count += 1
    
    await status_msg.edit_text(
        f"‚úÖ –†–∞—Å—Å—ã–ª–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞!\n\n"
        f"–£—Å–ø–µ—à–Ω–æ: {success_count}\n"
        f"–û—à–∏–±–æ–∫: {fail_count}"
    )


# ===== –£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ —Ä–∞—Å—Å—ã–ª–∫–∏ =====

@router.message(BroadcastStates.waiting_for_text)
async def process_broadcast_content(message: Message, state: FSMContext):
    """–£–Ω–∏–≤–µ—Ä—Å–∞–ª—å–Ω—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –¥–ª—è –ª—é–±–æ–≥–æ –∫–æ–Ω—Ç–µ–Ω—Ç–∞ —Ä–∞—Å—Å—ã–ª–∫–∏"""
    if not is_admin(message.from_user.id):
        return
    
    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –æ—Ç–º–µ–Ω—É
    if message.text == "/cancel":
        await state.clear()
        await message.answer("‚ùå –†–∞—Å—Å—ã–ª–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞", reply_markup=get_admin_keyboard())
        return
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="‚úÖ –û—Ç–ø—Ä–∞–≤–∏—Ç—å —Ä–∞—Å—Å—ã–ª–∫—É", callback_data="confirm_broadcast")],
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∏—Ç—å", callback_data="admin_broadcast")]
    ])
    
    # –û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–æ—Ç–æ
    if message.photo:
        photo_id = message.photo[-1].file_id
        caption = message.caption or ""
        
        await state.update_data(photo_id=photo_id, caption=caption, media_type="photo")
        
        preview_text = f"üñº <b>–ü—Ä–µ–≤—å—é —Ä–∞—Å—Å—ã–ª–∫–∏</b>\n\n"
        if caption:
            preview_text += f"{caption}\n\n"
        preview_text += "–ü–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç–µ –æ—Ç–ø—Ä–∞–≤–∫—É:"
        
        await message.answer_photo(
            photo=photo_id,
            caption=preview_text,
            reply_markup=keyboard
        )
    
    # –û–±—Ä–∞–±–æ—Ç–∫–∞ –≤–∏–¥–µ–æ
    elif message.video:
        video_id = message.video.file_id
        caption = message.caption or ""
        
        await state.update_data(video_id=video_id, caption=caption, media_type="video")
        
        preview_text = f"üé• <b>–ü—Ä–µ–≤—å—é —Ä–∞—Å—Å—ã–ª–∫–∏</b>\n\n"
        if caption:
            preview_text += f"{caption}\n\n"
        preview_text += "–ü–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç–µ –æ—Ç–ø—Ä–∞–≤–∫—É:"
        
        await message.answer_video(
            video=video_id,
            caption=preview_text,
            reply_markup=keyboard
        )
    
    # –û–±—Ä–∞–±–æ—Ç–∫–∞ —Ç–µ–∫—Å—Ç–∞
    elif message.text:
        text = message.text
        
        await state.update_data(text=text, media_type="text")
        
        await message.answer(
            f"üìù <b>–ü—Ä–µ–≤—å—é —Ä–∞—Å—Å—ã–ª–∫–∏:</b>\n\n{text}\n\n"
            "–ü–æ–¥—Ç–≤–µ—Ä–¥–∏—Ç–µ –æ—Ç–ø—Ä–∞–≤–∫—É:",
            reply_markup=keyboard
        )
    
    else:
        await message.answer("‚ùå –ü–æ–¥–¥–µ—Ä–∂–∏–≤–∞—é—Ç—Å—è —Ç–æ–ª—å–∫–æ —Ç–µ–∫—Å—Ç, —Ñ–æ—Ç–æ –∏–ª–∏ –≤–∏–¥–µ–æ")


@router.callback_query(F.data == "confirm_broadcast")
async def confirm_and_send_broadcast(callback: CallbackQuery, state: FSMContext):
    """–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ –∏ –æ—Ç–ø—Ä–∞–≤–∫–∞ —Ä–∞—Å—Å—ã–ª–∫–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    data = await state.get_data()
    media_type = data.get("media_type")
    
    db = await get_db()
    user_repo = UserRepository(db)
    
    users = await user_repo.get_all()
    
    success_count = 0
    fail_count = 0
    
    # –£–¥–∞–ª—è–µ–º –ø—Ä–µ–≤—å—é –∏ –æ—Ç–ø—Ä–∞–≤–ª—è–µ–º –Ω–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ –æ –Ω–∞—á–∞–ª–µ —Ä–∞—Å—Å—ã–ª–∫–∏
    try:
        await callback.message.delete()
    except:
        pass
    
    status_msg = await callback.bot.send_message(
        chat_id=callback.from_user.id,
        text=f"üì§ –ù–∞—á–∏–Ω–∞—é —Ä–∞—Å—Å—ã–ª–∫—É –¥–ª—è {len(users)} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π..."
    )
    
    for user in users:
        try:
            if media_type == "text":
                await callback.bot.send_message(
                    chat_id=user.telegram_id,
                    text=data["text"]
                )
            elif media_type == "photo":
                await callback.bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=data["photo_id"],
                    caption=data.get("caption")
                )
            elif media_type == "video":
                await callback.bot.send_video(
                    chat_id=user.telegram_id,
                    video=data["video_id"],
                    caption=data.get("caption")
                )
            
            success_count += 1
        except Exception:
            fail_count += 1
    
    await status_msg.edit_text(
        f"‚úÖ –†–∞—Å—Å—ã–ª–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞!\n\n"
        f"–£—Å–ø–µ—à–Ω–æ: {success_count}\n"
        f"–û—à–∏–±–æ–∫: {fail_count}"
    )
    
    await callback.bot.send_message(
        chat_id=callback.from_user.id,
        text="–í—ã–±–µ—Ä–∏—Ç–µ –¥–µ–π—Å—Ç–≤–∏–µ:",
        reply_markup=get_admin_keyboard()
    )
    
    await state.clear()
    await callback.answer()


@router.callback_query(F.data == "admin_courses")
async def show_courses_management(callback: CallbackQuery):
    """–ü–æ–∫–∞–∑–∞—Ç—å —É–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    # –ü–æ–ª—É—á–∞–µ–º –∫—É—Ä—Å—ã –∏–∑ JSON
    courses = get_all_courses()
    
    if not courses:
        await callback.message.edit_text(
            "üìö <b>–ö—É—Ä—Å—ã –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—Ç</b>\n\n"
            "–ö—É—Ä—Å—ã —Ö—Ä–∞–Ω—è—Ç—Å—è –≤ —Ñ–∞–π–ª–µ data/courses.json",
            reply_markup=get_back_to_admin_keyboard()
        )
    else:
        buttons = []
        for course in courses:
            buttons.append([InlineKeyboardButton(
                text=f"{course.get('emoji', 'üìö')} {course['name']}",
                callback_data=f"manage_course_{course['slug']}"
            )])
        
        buttons.append([InlineKeyboardButton(
            text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å",
            callback_data="admin_panel"
        )])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        
        await callback.message.edit_text(
            "üìö <b>–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏</b>\n\n"
            "–ü—Ä–æ—Å–º–æ—Ç—Ä –∫—É—Ä—Å–æ–≤. –î–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∏—Å–ø–æ–ª—å–∑—É–π—Ç–µ —Ñ–∞–π–ª—ã data/courses.json –∏ data/course_materials.json",
            reply_markup=keyboard
        )
    
    await callback.answer()


@router.callback_query(F.data.startswith("manage_course_"))
async def manage_course(callback: CallbackQuery):
    """–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–º –∫—É—Ä—Å–æ–º"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    # –ò–∑–≤–ª–µ–∫–∞–µ–º slug –∫—É—Ä—Å–∞
    course_slug = "_".join(callback.data.split("_")[2:])
    
    from data import get_course_by_slug, get_course_materials
    
    course = get_course_by_slug(course_slug)
    
    if not course:
        await callback.answer("–ö—É—Ä—Å –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    # –ü–æ–ª—É—á–∞–µ–º –º–∞—Ç–µ—Ä–∏–∞–ª—ã –∫—É—Ä—Å–∞
    materials = get_course_materials(course_slug)
    modules = materials.get('modules', []) if materials else []
    
    text = f"üìñ <b>{course['name']}</b>\n\n"
    text += f"üìç Slug: <code>{course['slug']}</code>\n"
    
    if course.get('price'):
        text += f"üí∞ –¶–µ–Ω–∞: {course['price']:,.0f} ‚ÇΩ\n"
    
    if course.get('duration'):
        text += f"‚è± –î–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å: {course['duration']}\n"
    
    text += f"‚úÖ –ê–∫—Ç–∏–≤–µ–Ω: {'–î–∞' if course.get('is_active', True) else '–ù–µ—Ç'}\n"
    
    if modules:
        text += f"\nüìÇ <b>–ú–æ–¥—É–ª–µ–π:</b> {len(modules)}\n\n"
        for module in modules:
            text += f"‚Ä¢ {module.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')}\n"
    else:
        text += "\n–ú–æ–¥—É–ª–µ–π –ø–æ–∫–∞ –Ω–µ—Ç\n"
    
    buttons = []
    
    # –ö–Ω–æ–ø–∫–∏ –¥–ª—è –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ –º–æ–¥—É–ª–µ–π
    if modules:
        buttons.append([InlineKeyboardButton(
            text="üìÇ –ü—Ä–æ—Å–º–æ—Ç—Ä –º–æ–¥—É–ª–µ–π",
            callback_data=f"view_modules_{course_slug}"
        )])
    
    buttons.append([InlineKeyboardButton(
        text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ –∫—É—Ä—Å–∞–º",
        callback_data="admin_courses"
    )])
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("view_modules_"))
async def view_modules(callback: CallbackQuery):
    """–ü—Ä–æ—Å–º–æ—Ç—Ä —Å–ø–∏—Å–∫–∞ –º–æ–¥—É–ª–µ–π –∫—É—Ä—Å–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    course_slug = "_".join(callback.data.split("_")[2:])
    
    from data import get_course_by_slug, get_course_modules
    
    course = get_course_by_slug(course_slug)
    if not course:
        await callback.answer("–ö—É—Ä—Å –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    modules = get_course_modules(course_slug)
    
    text = f"üìÇ <b>–ú–æ–¥—É–ª–∏ –∫—É—Ä—Å–∞: {course['name']}</b>\n\n"
    
    buttons = []
    for module in modules:
        buttons.append([InlineKeyboardButton(
            text=f"üìÇ {module.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')}",
            callback_data=f"view_module_{course_slug}_{module['id']}"
        )])
    
    # –ö–Ω–æ–ø–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –º–æ–¥—É–ª—è
    buttons.append([InlineKeyboardButton(
        text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –º–æ–¥—É–ª—å",
        callback_data=f"add_module_{course_slug}"
    )])
    
    buttons.append([InlineKeyboardButton(
        text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ –∫—É—Ä—Å—É",
        callback_data=f"manage_course_{course_slug}"
    )])
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("view_module_"))
async def view_module(callback: CallbackQuery):
    """–ü—Ä–æ—Å–º–æ—Ç—Ä –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–≥–æ –º–æ–¥—É–ª—è"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    module_id = parts[-1]
    course_slug = "_".join(parts[2:-1])
    
    from data import get_course_by_slug, get_module_by_id
    
    course = get_course_by_slug(course_slug)
    module = get_module_by_id(course_slug, module_id)
    
    if not course or not module:
        await callback.answer("–ú–æ–¥—É–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    text = f"üìÇ <b>{module['title']}</b>\n\n"
    if module.get('description'):
        text += f"{module['description']}\n\n"
    
    lessons = module.get('lessons', [])
    if lessons:
        text += f"üìù <b>–£—Ä–æ–∫–æ–≤: {len(lessons)}</b>\n\n"
        for i, lesson in enumerate(lessons, 1):
            text += f"{i}. {lesson.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')}\n"
    else:
        text += "–£—Ä–æ–∫–æ–≤ –ø–æ–∫–∞ –Ω–µ—Ç\n"
    
    buttons = []
    
    # –ö–Ω–æ–ø–∫–∏ –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —É—Ä–æ–∫–æ–≤
    if lessons:
        for lesson in lessons:
            buttons.append([InlineKeyboardButton(
                text=f"‚úèÔ∏è {lesson.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')[:30]}...",
                callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson['id']}"
            )])
    
    # –ö–Ω–æ–ø–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞
    buttons.append([InlineKeyboardButton(
        text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫",
        callback_data=f"add_lesson_to_module_{course_slug}_{module_id}"
    )])
    
    buttons.append([InlineKeyboardButton(
        text="‚óÄÔ∏è –ö –º–æ–¥—É–ª—è–º",
        callback_data=f"view_modules_{course_slug}"
    )])
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("add_module_"))
async def add_module_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –º–æ–¥—É–ª—è"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    course_slug = "_".join(callback.data.split("_")[2:])
    
    await state.update_data(course_slug=course_slug, adding_module=True)
    await state.set_state(CourseManagement.waiting_for_lesson_title)
    
    await callback.message.edit_text(
        "‚ûï <b>–î–æ–±–∞–≤–ª–µ–Ω–∏–µ –º–æ–¥—É–ª—è</b>\n\n"
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –º–æ–¥—É–ª—è:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"manage_course_{course_slug}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("add_lesson_to_module_"))
async def add_lesson_to_module_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞ –≤ –º–æ–¥—É–ª—å"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    module_id = parts[-1]
    course_slug = "_".join(parts[4:-1])
    
    await state.update_data(
        course_slug=course_slug,
        module_id=module_id,
        adding_lesson=True
    )
    await state.set_state(CourseManagement.waiting_for_lesson_title)
    
    await callback.message.edit_text(
        "‚ûï <b>–î–æ–±–∞–≤–ª–µ–Ω–∏–µ —É—Ä–æ–∫–∞</b>\n\n"
        "–í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
        ])
    )
    await callback.answer()


# –°—Ç–∞—Ä—ã–π –æ–±—Ä–∞–±–æ—Ç—á–∏–∫, –æ—Å—Ç–∞–≤–∏–º –¥–ª—è —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏, –Ω–æ –ø–µ—Ä–µ–Ω–∞–ø—Ä–∞–≤–∏–º
@router.callback_query(F.data.startswith("add_lesson_"))
async def add_lesson_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞ - DEPRECATED"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    # –ü–µ—Ä–µ–Ω–∞–ø—Ä–∞–≤–ª—è–µ–º –Ω–∞ —É–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏
    await callback.answer("‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –Ω–æ–≤—ã–π –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å", show_alert=True)
    await callback.message.edit_text(
        "‚ÑπÔ∏è –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏ –ø–µ—Ä–µ–Ω–µ—Å–µ–Ω–æ –Ω–∞ –Ω–æ–≤—ã–π –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å\n\n"
        "–ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –∫–Ω–æ–ø–∫—É '–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏' –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª–∏",
        reply_markup=get_back_to_admin_keyboard()
    )


@router.message(CourseManagement.waiting_for_lesson_title)
async def process_title_input(message: Message, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤–≤–æ–¥–∞ –Ω–∞–∑–≤–∞–Ω–∏—è (–º–æ–¥—É–ª—è –∏–ª–∏ —É—Ä–æ–∫–∞)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    title = message.text.strip()
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ –¥–æ–±–∞–≤–ª—è–µ–º
    if data.get('adding_module'):
        # –î–æ–±–∞–≤–ª—è–µ–º –º–æ–¥—É–ª—å
        await state.update_data(module_title=title)
        course_slug = data['course_slug']
        
        await message.answer(
            f"‚úÖ –ù–∞–∑–≤–∞–Ω–∏–µ –º–æ–¥—É–ª—è: <b>{title}</b>\n\n"
            f"–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ –æ–ø–∏—Å–∞–Ω–∏–µ –º–æ–¥—É–ª—è (–∏–ª–∏ –æ—Ç–ø—Ä–∞–≤—å—Ç–µ '–ø—Ä–æ–ø—É—Å—Ç–∏—Ç—å'):",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_module_desc_{course_slug}")]
            ])
        )
        await state.set_state(CourseManagement.waiting_for_lesson_description)
        
    elif data.get('adding_lesson'):
        # –î–æ–±–∞–≤–ª—è–µ–º —É—Ä–æ–∫
        await state.update_data(lesson_title=title)
        course_slug = data['course_slug']
        module_id = data['module_id']
        
        await message.answer(
            f"‚úÖ –ù–∞–∑–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞: <b>{title}</b>\n\n"
            f"–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ –æ–ø–∏—Å–∞–Ω–∏–µ —É—Ä–æ–∫–∞:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
            ])
        )
        await state.set_state(CourseManagement.waiting_for_lesson_description)


@router.callback_query(F.data.startswith("skip_module_desc_"))
async def skip_module_description(callback: CallbackQuery, state: FSMContext):
    """–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å –æ–ø–∏—Å–∞–Ω–∏–µ –º–æ–¥—É–ª—è"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await state.update_data(module_description="")
    await save_new_module(callback.message, state)
    await callback.answer()


@router.message(CourseManagement.waiting_for_lesson_description)
async def process_description_input(message: Message, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤–≤–æ–¥–∞ –æ–ø–∏—Å–∞–Ω–∏—è (–º–æ–¥—É–ª—è –∏–ª–∏ —É—Ä–æ–∫–∞)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    description = message.text.strip()
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ –¥–æ–±–∞–≤–ª—è–µ–º
    if data.get('adding_module'):
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –æ–ø–∏—Å–∞–Ω–∏–µ –∏ —Å–æ–∑–¥–∞–µ–º –º–æ–¥—É–ª—å
        await state.update_data(module_description=description)
        await save_new_module(message, state)
        
    elif data.get('adding_lesson'):
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –æ–ø–∏—Å–∞–Ω–∏–µ –∏ –ø—Ä–æ—Å–∏–º –≤–∏–¥–µ–æ
        await state.update_data(lesson_description=description)
        course_slug = data['course_slug']
        module_id = data['module_id']
        
        await message.answer(
            "‚úÖ –û–ø–∏—Å–∞–Ω–∏–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–æ\n\n"
            "–û—Ç–ø—Ä–∞–≤—å—Ç–µ —Å—Å—ã–ª–∫—É –Ω–∞ –≤–∏–¥–µ–æ (YouTube URL) –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å':",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_lesson_video_{course_slug}_{module_id}")],
                [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
            ])
        )
        await state.set_state(CourseManagement.waiting_for_lesson_video)


async def save_new_module(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –º–æ–¥—É–ª—è"""
    data = await state.get_data()
    course_slug = data['course_slug']
    
    from data import add_module_to_course, get_course_modules
    import uuid
    
    # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º ID
    modules = get_course_modules(course_slug)
    module_order = len(modules) + 1
    
    new_module = {
        'id': f"module-{module_order}",
        'title': data.get('module_title', '–ù–æ–≤—ã–π –º–æ–¥—É–ª—å'),
        'description': data.get('module_description', ''),
        'order': module_order,
        'lessons': []
    }
    
    try:
        add_module_to_course(course_slug, new_module)
        
        await message.answer(
            f"‚úÖ <b>–ú–æ–¥—É–ª—å —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω!</b>\n\n"
            f"üìÇ {new_module['title']}\n\n"
            f"–¢–µ–ø–µ—Ä—å –≤—ã –º–æ–∂–µ—Ç–µ –¥–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫–∏ –≤ —ç—Ç–æ—Ç –º–æ–¥—É–ª—å.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫", callback_data=f"add_lesson_to_module_{course_slug}_{new_module['id']}")],
                [InlineKeyboardButton(text="‚óÄÔ∏è –ö –∫—É—Ä—Å—É", callback_data=f"manage_course_{course_slug}")]
            ])
        )
    except Exception as e:
        await message.answer(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ–±–∞–≤–ª–µ–Ω–∏–∏ –º–æ–¥—É–ª—è: {str(e)}")
    
    await state.clear()


@router.callback_query(F.data.startswith("skip_lesson_video_"))
async def skip_lesson_video(callback: CallbackQuery, state: FSMContext):
    """–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å –≤–∏–¥–µ–æ –∏ –ø–µ—Ä–µ–π—Ç–∏ –∫ –ª–µ–∫—Ü–∏–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await state.update_data(video_url="")
    
    # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ –∑–∞–ø—Ä–æ—Å—É PDF –ª–µ–∫—Ü–∏–∏
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    
    await callback.message.edit_text(
        "–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ PDF —Ñ–∞–π–ª –ª–µ–∫—Ü–∏–∏ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å':",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_lesson_lecture_{course_slug}_{module_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
        ])
    )
    await state.set_state(CourseManagement.waiting_for_lesson_lecture_file)
    await callback.answer()


@router.callback_query(F.data.startswith("skip_lesson_lecture_"))
async def skip_lesson_lecture(callback: CallbackQuery, state: FSMContext):
    """–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å PDF –ª–µ–∫—Ü–∏—é –∏ –ø–µ—Ä–µ–π—Ç–∏ –∫ —Ç–µ–∫—Å—Ç–æ–≤–æ–º—É —Å–æ–¥–µ—Ä–∂–∏–º–æ–º—É"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await state.update_data(lecture_file_id="")
    
    # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ –∑–∞–ø—Ä–æ—Å—É —Ç–µ–∫—Å—Ç–æ–≤–æ–≥–æ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    
    await callback.message.edit_text(
        "–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ —Ç–µ–∫—Å—Ç–æ–≤–æ–µ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å':",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_lesson_text_{course_slug}_{module_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
        ])
    )
    await state.set_state(CourseManagement.waiting_for_lesson_text_content)
    await callback.answer()


@router.callback_query(F.data.startswith("skip_lesson_text_"))
async def skip_lesson_text(callback: CallbackQuery, state: FSMContext):
    """–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å —Ç–µ–∫—Å—Ç–æ–≤–æ–µ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏–µ –∏ —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å —É—Ä–æ–∫"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await state.update_data(text_content="")
    await save_new_lesson_to_module(callback.message, state)
    await callback.answer()


async def process_lesson_lecture(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ PDF –ª–µ–∫—Ü–∏–∏ –¥–ª—è —É—Ä–æ–∫–∞ (–≤—Å–ø–æ–º–æ–≥–∞—Ç–µ–ª—å–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –Ω–æ–≤–æ–≥–æ —É—Ä–æ–∫–∞)"""
    if not is_admin(message.from_user.id):
        return
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ —ç—Ç–æ PDF
    if not message.document.file_name.lower().endswith('.pdf'):
        await message.answer("‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ PDF —Ñ–∞–π–ª")
        return
    
    lecture_file_id = message.document.file_id
    await state.update_data(lecture_file_id=lecture_file_id)
    
    # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ –∑–∞–ø—Ä–æ—Å—É —Ç–µ–∫—Å—Ç–æ–≤–æ–≥–æ —Å–æ–¥–µ—Ä–∂–∏–º–æ–≥–æ
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    
    await message.answer(
        "‚úÖ PDF –ª–µ–∫—Ü–∏—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∞\n\n"
        "–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ —Ç–µ–∫—Å—Ç–æ–≤–æ–µ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å':",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_lesson_text_{course_slug}_{module_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
        ])
    )
    await state.set_state(CourseManagement.waiting_for_lesson_text_content)


# –≠—Ç–æ—Ç –æ–±—Ä–∞–±–æ—Ç—á–∏–∫ –±–æ–ª—å—à–µ –Ω–µ –Ω—É–∂–µ–Ω - –ª–æ–≥–∏–∫–∞ –ø–µ—Ä–µ–Ω–µ—Å–µ–Ω–∞ –≤ save_json_lesson_lecture
# async def invalid_lecture_file(message: Message):
#     """–û–±—Ä–∞–±–æ—Ç–∫–∞ –Ω–µ–≤–µ—Ä–Ω–æ–≥–æ —Ñ–æ—Ä–º–∞—Ç–∞ —Ñ–∞–π–ª–∞ –ª–µ–∫—Ü–∏–∏"""
#     if not is_admin(message.from_user.id):
#         return
#     
#     await message.answer(
#         "‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ PDF —Ñ–∞–π–ª (–Ω–µ —Å—Å—ã–ª–∫—É, –Ω–µ –∏–∑–æ–±—Ä–∞–∂–µ–Ω–∏–µ)\n\n"
#         "–ò–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å' –¥–ª—è –ø—Ä–æ–¥–æ–ª–∂–µ–Ω–∏—è"
#     )


@router.message(CourseManagement.waiting_for_lesson_video)
async def process_lesson_video(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å—Å—ã–ª–∫–∏ –Ω–∞ –≤–∏–¥–µ–æ –¥–ª—è —É—Ä–æ–∫–∞"""
    if not is_admin(message.from_user.id):
        return
    
    video_url = message.text.strip()
    await state.update_data(video_url=video_url)
    
    # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ –∑–∞–ø—Ä–æ—Å—É PDF –ª–µ–∫—Ü–∏–∏
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    
    await message.answer(
        "‚úÖ –°—Å—ã–ª–∫–∞ –Ω–∞ –≤–∏–¥–µ–æ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∞\n\n"
        "–¢–µ–ø–µ—Ä—å –æ—Ç–ø—Ä–∞–≤—å—Ç–µ PDF —Ñ–∞–π–ª –ª–µ–∫—Ü–∏–∏ –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å':",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚è≠ –ü—Ä–æ–ø—É—Å—Ç–∏—Ç—å", callback_data=f"skip_lesson_lecture_{course_slug}_{module_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"view_module_{course_slug}_{module_id}")]
        ])
    )
    await state.set_state(CourseManagement.waiting_for_lesson_lecture_file)


@router.message(CourseManagement.waiting_for_lesson_text_content)
async def process_lesson_text_content(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ —Ç–µ–∫—Å—Ç–æ–≤–æ–≥–æ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏—è —É—Ä–æ–∫–∞"""
    if not is_admin(message.from_user.id):
        return
    
    text_content = message.text.strip()
    await state.update_data(text_content=text_content)
    await save_new_lesson_to_module(message, state)


async def save_new_lesson_to_module(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ —É—Ä–æ–∫–∞ –≤ –º–æ–¥—É–ª—å"""
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    
    from data import add_lesson_to_module, get_module_by_id
    
    # –ü–æ–ª—É—á–∞–µ–º –º–æ–¥—É–ª—å —á—Ç–æ–±—ã –æ–ø—Ä–µ–¥–µ–ª–∏—Ç—å –ø–æ—Ä—è–¥–æ–∫ —É—Ä–æ–∫–∞
    module = get_module_by_id(course_slug, module_id)
    if not module:
        await message.answer("‚ùå –ú–æ–¥—É–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω")
        await state.clear()
        return
    
    lessons_count = len(module.get('lessons', []))
    lesson_order = lessons_count + 1
    
    new_lesson = {
        'id': f"lesson-{module_id}-{lesson_order}",
        'title': data.get('lesson_title', '–ù–æ–≤—ã–π —É—Ä–æ–∫'),
        'description': data.get('lesson_description', ''),
        'duration': '',
        'type': 'video',
        'video_url': data.get('video_url', ''),
        'lecture_file_id': data.get('lecture_file_id', ''),
        'text_content': data.get('text_content', ''),
        'order': lesson_order,
        'materials': []
    }
    
    try:
        success = add_lesson_to_module(course_slug, module_id, new_lesson)
        
        if success:
            # –§–æ—Ä–º–∏—Ä—É–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –¥–æ–±–∞–≤–ª–µ–Ω–Ω–æ–º –∫–æ–Ω—Ç–µ–Ω—Ç–µ
            content_info = []
            if new_lesson['video_url']:
                content_info.append("üé¨ –í–∏–¥–µ–æ")
            if new_lesson['lecture_file_id']:
                content_info.append("üìÑ PDF –ª–µ–∫—Ü–∏—è")
            if new_lesson['text_content']:
                content_info.append("üìù –¢–µ–∫—Å—Ç")
            
            content_text = ", ".join(content_info) if content_info else "–ë–µ–∑ –∫–æ–Ω—Ç–µ–Ω—Ç–∞"
            
            await message.answer(
                f"‚úÖ <b>–£—Ä–æ–∫ —É—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω!</b>\n\n"
                f"üìù {new_lesson['title']}\n"
                f"üìÇ –ú–æ–¥—É–ª—å: {module['title']}\n"
                f"üì¶ –ö–æ–Ω—Ç–µ–Ω—Ç: {content_text}",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å –µ—â–µ —É—Ä–æ–∫", callback_data=f"add_lesson_to_module_{course_slug}_{module_id}")],
                    [InlineKeyboardButton(text="‚óÄÔ∏è –ö –º–æ–¥—É–ª—é", callback_data=f"view_module_{course_slug}_{module_id}")]
                ])
            )
        else:
            await message.answer("‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ–±–∞–≤–ª–µ–Ω–∏–∏ —É—Ä–æ–∫–∞")
    except Exception as e:
        await message.answer(f"‚ùå –û—à–∏–±–∫–∞: {str(e)}")
    
    await state.clear()


# DEPRECATED: –°—Ç–∞—Ä—ã–µ –æ–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —É—Ä–æ–∫–æ–≤ (—Å –ë–î) –∑–∞–∫–æ–º–º–µ–Ω—Ç–∏—Ä–æ–≤–∞–Ω—ã

@router.callback_query(F.data.startswith("edit_lessons_"))
async def edit_lessons_list(callback: CallbackQuery):
    """–°–ø–∏—Å–æ–∫ —É—Ä–æ–∫–æ–≤ –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è - DEPRECATED"""
    await callback.answer("‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –Ω–æ–≤—ã–π –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å —á–µ—Ä–µ–∑ '–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏'", show_alert=True)
    return
    
    # –°–¢–ê–†–´–ô –ö–û–î –ó–ê–ö–û–ú–ú–ï–ù–¢–ò–†–û–í–ê–ù
    course_id = int(callback.data.split("_")[2])
    db: Session = get_db()
    
    try:
        course = db.query(Course).filter(Course.id == course_id).first()
        lessons = db.query(Lesson).filter(Lesson.course_id == course_id).order_by(
            Lesson.module_number, Lesson.lesson_number
        ).all()
        
        buttons = []
        for lesson in lessons:
            buttons.append([InlineKeyboardButton(
                text=f"–ú{lesson.module_number}.–£{lesson.lesson_number}: {lesson.title}",
                callback_data=f"edit_lesson_{lesson.id}"
            )])
        
        buttons.append([InlineKeyboardButton(
            text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ –∫—É—Ä—Å—É",
            callback_data=f"manage_course_{course_id}"
        )])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        
        await callback.message.edit_text(
            f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —É—Ä–æ–∫–æ–≤ –∫—É—Ä—Å–∞ \"{course.name}\"</b>\n\n"
            f"–í—ã–±–µ—Ä–∏—Ç–µ —É—Ä–æ–∫:",
            reply_markup=keyboard
        )
    
    finally:
        db.close()
    
    await callback.answer()


@router.callback_query(F.data.startswith("edit_lesson_"))
async def edit_lesson_menu(callback: CallbackQuery):
    """–ú–µ–Ω—é —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —É—Ä–æ–∫–∞ (–Ω–æ–≤—ã–π —Ñ–æ—Ä–º–∞—Ç —Å JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    # –ù–æ–≤—ã–π —Ñ–æ—Ä–º–∞—Ç: edit_lesson_{course_slug}_{module_id}_{lesson_id}
    parts = callback.data.split("_")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ñ–æ—Ä–º–∞—Ç - –µ—Å–ª–∏ —Å—Ç–∞—Ä—ã–π (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ), –ø–æ–∫–∞–∑—ã–≤–∞–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ
    if len(parts) == 3:
        try:
            int(parts[2])  # –ü—ã—Ç–∞–µ–º—Å—è –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞—Ç—å –≤ —á–∏—Å–ª–æ
            await callback.answer("‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –Ω–æ–≤—ã–π –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å —á–µ—Ä–µ–∑ '–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏'", show_alert=True)
            return
        except ValueError:
            pass
    
    # –û–±—Ä–∞–±–æ—Ç–∫–∞ –Ω–æ–≤–æ–≥–æ —Ñ–æ—Ä–º–∞—Ç–∞
    if len(parts) < 5:
        await callback.answer("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç –¥–∞–Ω–Ω—ã—Ö", show_alert=True)
        return
    
    course_slug = parts[2]
    module_id = parts[3]
    lesson_id = parts[4]
    
    from data import get_lesson_by_id
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    
    if not lesson:
        await callback.answer("–£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    text = f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞</b>\n\n"
    text += f"üìù –ù–∞–∑–≤–∞–Ω–∏–µ: {lesson['title']}\n"
    text += f"üìÑ –û–ø–∏—Å–∞–Ω–∏–µ: {lesson.get('description', '–ù–µ —É–∫–∞–∑–∞–Ω–æ')[:100]}...\n\n"
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –∫–æ–Ω—Ç–µ–Ω—Ç–∞
    has_video = bool(lesson.get('video_url'))
    has_lecture = bool(lesson.get('lecture_file_id'))
    has_text = bool(lesson.get('text_content'))
    
    text += f"üé¨ –í–∏–¥–µ–æ: {'‚úÖ –ï—Å—Ç—å' if has_video else '‚ùå –ù–µ—Ç'}\n"
    text += f"üìÑ PDF –ª–µ–∫—Ü–∏—è: {'‚úÖ –ï—Å—Ç—å' if has_lecture else '‚ùå –ù–µ—Ç'}\n"
    text += f"üìù –¢–µ–∫—Å—Ç: {'‚úÖ –ï—Å—Ç—å' if has_text else '‚ùå –ù–µ—Ç'}\n"
    
    buttons = [
        [InlineKeyboardButton(text="‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –Ω–∞–∑–≤–∞–Ω–∏–µ", callback_data=f"edit_json_title_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –æ–ø–∏—Å–∞–Ω–∏–µ", callback_data=f"edit_json_desc_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="üé¨ –ò–∑–º–µ–Ω–∏—Ç—å –≤–∏–¥–µ–æ", callback_data=f"edit_json_video_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="üìÑ –ò–∑–º–µ–Ω–∏—Ç—å PDF –ª–µ–∫—Ü–∏—é", callback_data=f"edit_json_lecture_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="üìù –ò–∑–º–µ–Ω–∏—Ç—å —Ç–µ–∫—Å—Ç", callback_data=f"edit_json_text_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="üóë –£–¥–∞–ª–∏—Ç—å —É—Ä–æ–∫", callback_data=f"delete_json_lesson_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="‚óÄÔ∏è –ö –º–æ–¥—É–ª—é", callback_data=f"view_module_{course_slug}_{module_id}")]
    ]
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


# ===== –†–ï–î–ê–ö–¢–ò–†–û–í–ê–ù–ò–ï –£–†–û–ö–û–í (JSON –§–û–†–ú–ê–¢) =====

@router.callback_query(F.data.startswith("edit_json_title_"))
async def edit_json_lesson_title_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –Ω–∞–∑–≤–∞–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    await state.update_data(course_slug=course_slug, module_id=module_id, lesson_id=lesson_id)
    await state.set_state(CourseManagement.editing_lesson_title)
    
    await callback.message.edit_text(
        "‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –Ω–∞–∑–≤–∞–Ω–∏—è —É—Ä–æ–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("edit_json_desc_"))
async def edit_json_lesson_desc_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –æ–ø–∏—Å–∞–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    await state.update_data(course_slug=course_slug, module_id=module_id, lesson_id=lesson_id)
    await state.set_state(CourseManagement.editing_lesson_description)
    
    await callback.message.edit_text(
        "‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –æ–ø–∏—Å–∞–Ω–∏—è —É—Ä–æ–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤–æ–µ –æ–ø–∏—Å–∞–Ω–∏–µ:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("edit_json_video_"))
async def edit_json_lesson_video_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –≤–∏–¥–µ–æ —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    await state.update_data(course_slug=course_slug, module_id=module_id, lesson_id=lesson_id)
    await state.set_state(CourseManagement.editing_lesson_video)
    
    await callback.message.edit_text(
        "üé¨ <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –≤–∏–¥–µ–æ —É—Ä–æ–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤—É—é —Å—Å—ã–ª–∫—É –Ω–∞ –≤–∏–¥–µ–æ –∏–ª–∏ '-' –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("edit_json_lecture_"))
async def edit_json_lesson_lecture_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è PDF –ª–µ–∫—Ü–∏–∏ —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    # –û—á–∏—â–∞–µ–º state –∏ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
    await state.clear()
    await state.update_data(
        course_slug=course_slug, 
        module_id=module_id, 
        lesson_id=lesson_id, 
        editing_lecture=True,
        adding_lesson=False  # —è–≤–Ω–æ —É–∫–∞–∑—ã–≤–∞–µ–º —á—Ç–æ —ç—Ç–æ –ù–ï –¥–æ–±–∞–≤–ª–µ–Ω–∏–µ
    )
    await state.set_state(CourseManagement.waiting_for_lesson_lecture_file)
    
    print(f"üéØ –ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è PDF –ª–µ–∫—Ü–∏–∏ –¥–ª—è —É—Ä–æ–∫–∞: {lesson_id}")
    
    await callback.message.edit_text(
        "üìÑ <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ PDF –ª–µ–∫—Ü–∏–∏ —É—Ä–æ–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤—ã–π PDF —Ñ–∞–π–ª –∏–ª–∏ –Ω–∞–∂–º–∏—Ç–µ '–£–¥–∞–ª–∏—Ç—å' –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–π –ª–µ–∫—Ü–∏–∏:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="üóë –£–¥–∞–ª–∏—Ç—å –ª–µ–∫—Ü–∏—é", callback_data=f"remove_lecture_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("remove_lecture_"))
async def remove_json_lesson_lecture(callback: CallbackQuery, state: FSMContext):
    """–£–¥–∞–ª–µ–Ω–∏–µ PDF –ª–µ–∫—Ü–∏–∏ –∏–∑ —É—Ä–æ–∫–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[2]
    module_id = parts[3]
    lesson_id = parts[4]
    
    from data import get_lesson_by_id, update_lesson
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    if lesson:
        lesson['lecture_file_id'] = ""
        update_lesson(course_slug, module_id, lesson_id, lesson)
        
        await callback.answer("‚úÖ PDF –ª–µ–∫—Ü–∏—è —É–¥–∞–ª–µ–Ω–∞", show_alert=True)
        await state.clear()
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º—Å—è –∫ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—é —É—Ä–æ–∫–∞ - –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –º–µ–Ω—é –∑–∞–Ω–æ–≤–æ
        text = f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞</b>\n\n"
        text += f"üìù –ù–∞–∑–≤–∞–Ω–∏–µ: {lesson['title']}\n"
        text += f"üìÑ –û–ø–∏—Å–∞–Ω–∏–µ: {lesson.get('description', '–ù–µ —É–∫–∞–∑–∞–Ω–æ')[:100]}...\n\n"
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –∫–æ–Ω—Ç–µ–Ω—Ç–∞
        has_video = bool(lesson.get('video_url'))
        has_lecture = False  # –ú—ã —Ç–æ–ª—å–∫–æ —á—Ç–æ —É–¥–∞–ª–∏–ª–∏
        has_text = bool(lesson.get('text_content'))
        
        text += f"üé¨ –í–∏–¥–µ–æ: {'‚úÖ –ï—Å—Ç—å' if has_video else '‚ùå –ù–µ—Ç'}\n"
        text += f"üìÑ PDF –ª–µ–∫—Ü–∏—è: ‚ùå –ù–µ—Ç\n"
        text += f"üìù –¢–µ–∫—Å—Ç: {'‚úÖ –ï—Å—Ç—å' if has_text else '‚ùå –ù–µ—Ç'}\n"
        
        buttons = [
            [InlineKeyboardButton(text="‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –Ω–∞–∑–≤–∞–Ω–∏–µ", callback_data=f"edit_json_title_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –æ–ø–∏—Å–∞–Ω–∏–µ", callback_data=f"edit_json_desc_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="üé¨ –ò–∑–º–µ–Ω–∏—Ç—å –≤–∏–¥–µ–æ", callback_data=f"edit_json_video_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="üìÑ –ò–∑–º–µ–Ω–∏—Ç—å PDF –ª–µ–∫—Ü–∏—é", callback_data=f"edit_json_lecture_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="üìù –ò–∑–º–µ–Ω–∏—Ç—å —Ç–µ–∫—Å—Ç", callback_data=f"edit_json_text_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="üóë –£–¥–∞–ª–∏—Ç—å —É—Ä–æ–∫", callback_data=f"delete_json_lesson_{course_slug}_{module_id}_{lesson_id}")],
            [InlineKeyboardButton(text="‚óÄÔ∏è –ö –º–æ–¥—É–ª—é", callback_data=f"view_module_{course_slug}_{module_id}")]
        ]
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        await callback.message.edit_text(text, reply_markup=keyboard)
    else:
        await callback.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)


@router.callback_query(F.data.startswith("edit_json_text_"))
async def edit_json_lesson_text_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —Ç–µ–∫—Å—Ç–∞ —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    await state.update_data(course_slug=course_slug, module_id=module_id, lesson_id=lesson_id)
    await state.set_state(CourseManagement.editing_lesson_content)
    
    await callback.message.edit_text(
        "üìù <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–æ–≤–æ–≥–æ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏—è —É—Ä–æ–∫–∞</b>\n\n"
        "–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤—ã–π —Ç–µ–∫—Å—Ç –∏–ª–∏ '-' –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
        ])
    )
    await callback.answer()


@router.callback_query(F.data.startswith("delete_json_lesson_"))
async def delete_json_lesson_confirm(callback: CallbackQuery):
    """–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    from data import get_lesson_by_id
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    
    if not lesson:
        await callback.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    text = f"‚ö†Ô∏è <b>–£–¥–∞–ª–µ–Ω–∏–µ —É—Ä–æ–∫–∞</b>\n\n"
    text += f"–í—ã —É–≤–µ—Ä–µ–Ω—ã, —á—Ç–æ —Ö–æ—Ç–∏—Ç–µ —É–¥–∞–ª–∏—Ç—å —É—Ä–æ–∫:\n"
    text += f"<b>{lesson['title']}</b>?\n\n"
    text += "–≠—Ç–æ –¥–µ–π—Å—Ç–≤–∏–µ –Ω–µ–ª—å–∑—è –æ—Ç–º–µ–Ω–∏—Ç—å!"
    
    buttons = [
        [InlineKeyboardButton(text="‚úÖ –î–∞, —É–¥–∞–ª–∏—Ç—å", callback_data=f"confirm_delete_json_{course_slug}_{module_id}_{lesson_id}")],
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
    ]
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("confirm_delete_json_"))
async def delete_json_lesson_execute(callback: CallbackQuery):
    """–í—ã–ø–æ–ª–Ω–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    parts = callback.data.split("_")
    course_slug = parts[3]
    module_id = parts[4]
    lesson_id = parts[5]
    
    from data import delete_lesson
    
    success = delete_lesson(course_slug, module_id, lesson_id)
    
    if success:
        await callback.answer("‚úÖ –£—Ä–æ–∫ —É–¥–∞–ª–µ–Ω", show_alert=True)
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º—Å—è –∫ –º–æ–¥—É–ª—é - –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –µ–≥–æ –∑–∞–Ω–æ–≤–æ
        from data import get_module_by_id, get_course_by_slug
        
        module = get_module_by_id(course_slug, module_id)
        course = get_course_by_slug(course_slug)
        
        if not module or not course:
            await callback.message.edit_text(
                "‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ –º–æ–¥—É–ª—è",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥", callback_data=f"manage_course_{course_slug}")]
                ])
            )
            return
        
        text = f"üìÇ <b>{module['title']}</b>\n\n"
        if module.get('description'):
            text += f"{module['description']}\n\n"
        
        lessons = module.get('lessons', [])
        if lessons:
            text += f"üìù <b>–£—Ä–æ–∫–æ–≤: {len(lessons)}</b>\n\n"
            for i, lesson in enumerate(lessons, 1):
                text += f"{i}. {lesson.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')}\n"
        else:
            text += "–£—Ä–æ–∫–æ–≤ –ø–æ–∫–∞ –Ω–µ—Ç\n"
        
        buttons = []
        
        # –ö–Ω–æ–ø–∫–∏ –¥–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è —É—Ä–æ–∫–æ–≤
        if lessons:
            for lesson in lessons:
                buttons.append([InlineKeyboardButton(
                    text=f"‚úèÔ∏è {lesson.get('title', '–ë–µ–∑ –Ω–∞–∑–≤–∞–Ω–∏—è')[:30]}...",
                    callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson['id']}"
                )])
        
        # –ö–Ω–æ–ø–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞
        buttons.append([InlineKeyboardButton(
            text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫",
            callback_data=f"add_lesson_to_module_{course_slug}_{module_id}"
        )])
        
        buttons.append([InlineKeyboardButton(
            text="‚óÄÔ∏è –ö –º–æ–¥—É–ª—è–º",
            callback_data=f"view_modules_{course_slug}"
        )])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        await callback.message.edit_text(text, reply_markup=keyboard)
    else:
        await callback.answer("‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏ —É—Ä–æ–∫–∞", show_alert=True)


# ===== –û–ë–†–ê–ë–û–¢–ß–ò–ö–ò –°–û–•–†–ê–ù–ï–ù–ò–Ø –ò–ó–ú–ï–ù–ï–ù–ò–ô =====

@router.message(CourseManagement.editing_lesson_title)
async def save_json_lesson_title(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –Ω–∞–∑–≤–∞–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    lesson_id = data['lesson_id']
    
    from data import get_lesson_by_id, update_lesson
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    if lesson:
        lesson['title'] = message.text.strip()
        update_lesson(course_slug, module_id, lesson_id, lesson)
        
        await message.answer(
            f"‚úÖ –ù–∞–∑–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–æ!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚óÄÔ∏è –ö —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
            ])
        )
    else:
        await message.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω")
    
    await state.clear()


@router.message(CourseManagement.editing_lesson_description)
async def save_json_lesson_description(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –æ–ø–∏—Å–∞–Ω–∏—è —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    lesson_id = data['lesson_id']
    
    from data import get_lesson_by_id, update_lesson
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    if lesson:
        lesson['description'] = message.text.strip()
        update_lesson(course_slug, module_id, lesson_id, lesson)
        
        await message.answer(
            f"‚úÖ –û–ø–∏—Å–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–æ!",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚óÄÔ∏è –ö —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
            ])
        )
    else:
        await message.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω")
    
    await state.clear()


@router.message(CourseManagement.editing_lesson_video)
async def save_json_lesson_video(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–π —Å—Å—ã–ª–∫–∏ –Ω–∞ –≤–∏–¥–µ–æ —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    lesson_id = data['lesson_id']
    
    from data import get_lesson_by_id, update_lesson
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    if lesson:
        video_url = message.text.strip()
        lesson['video_url'] = "" if video_url == "-" else video_url
        update_lesson(course_slug, module_id, lesson_id, lesson)
        
        if video_url == "-":
            msg = "‚úÖ –í–∏–¥–µ–æ —É–¥–∞–ª–µ–Ω–æ –∏–∑ —É—Ä–æ–∫–∞!"
        else:
            msg = "‚úÖ –°—Å—ã–ª–∫–∞ –Ω–∞ –≤–∏–¥–µ–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∞!"
        
        await message.answer(
            msg,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚óÄÔ∏è –ö —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
            ])
        )
    else:
        await message.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω")
    
    await state.clear()


@router.message(CourseManagement.waiting_for_lesson_lecture_file, F.document)
async def save_json_lesson_lecture(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ PDF –ª–µ–∫—Ü–∏–∏ —É—Ä–æ–∫–∞ (—Å–æ–∑–¥–∞–Ω–∏–µ –∏–ª–∏ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    
    # –û–¢–õ–ê–î–ö–ê: –ª–æ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ
    logger.info(f"üîç PDF Upload - State data: editing_lecture={data.get('editing_lecture')}, lesson_id={data.get('lesson_id')}, adding_lesson={data.get('adding_lesson')}")
    print(f"üîç PDF Upload - State data: editing_lecture={data.get('editing_lecture')}, lesson_id={data.get('lesson_id')}, adding_lesson={data.get('adding_lesson')}")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ —ç—Ç–æ PDF
    if not message.document.file_name.lower().endswith('.pdf'):
        await message.answer("‚ùå –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –æ—Ç–ø—Ä–∞–≤—å—Ç–µ PDF —Ñ–∞–π–ª")
        return
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —ç—Ç–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —É—Ä–æ–∫–∞, –∞ –Ω–µ —Å–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ
    # –†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ: –µ—Å—Ç—å editing_lecture –ò –µ—Å—Ç—å lesson_id –ò –ù–ï–¢ adding_lesson
    if data.get('editing_lecture') and data.get('lesson_id') and not data.get('adding_lesson'):
        # –≠–¢–û –†–ï–î–ê–ö–¢–ò–†–û–í–ê–ù–ò–ï –°–£–©–ï–°–¢–í–£–Æ–©–ï–ì–û –£–†–û–ö–ê
        logger.info(f"‚úÖ PDF Upload - –†–ï–î–ê–ö–¢–ò–†–û–í–ê–ù–ò–ï —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —É—Ä–æ–∫–∞: {data.get('lesson_id')}")
        print(f"‚úÖ PDF Upload - –†–ï–î–ê–ö–¢–ò–†–û–í–ê–ù–ò–ï —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ —É—Ä–æ–∫–∞: {data.get('lesson_id')}")
        course_slug = data['course_slug']
        module_id = data['module_id']
        lesson_id = data['lesson_id']
        
        from data import get_lesson_by_id, update_lesson
        
        lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
        if lesson:
            lesson['lecture_file_id'] = message.document.file_id
            update_lesson(course_slug, module_id, lesson_id, lesson)
            
            await message.answer(
                "‚úÖ PDF –ª–µ–∫—Ü–∏—è –æ–±–Ω–æ–≤–ª–µ–Ω–∞!",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="‚óÄÔ∏è –ö —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
                ])
            )
        else:
            await message.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω")
        
        await state.clear()
        return
    
    # –≠—Ç–æ —Å–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ —É—Ä–æ–∫–∞ - –ø—Ä–æ–¥–æ–ª–∂–∞–µ–º –ø—Ä–æ—Ü–µ—Å—Å
    logger.info(f"‚ûï PDF Upload - –°–û–ó–î–ê–ù–ò–ï –Ω–æ–≤–æ–≥–æ —É—Ä–æ–∫–∞")
    print(f"‚ûï PDF Upload - –°–û–ó–î–ê–ù–ò–ï –Ω–æ–≤–æ–≥–æ —É—Ä–æ–∫–∞")
    await process_lesson_lecture(message, state)


@router.message(CourseManagement.editing_lesson_content)
async def save_json_lesson_text(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ —Ç–µ–∫—Å—Ç–∞ —É—Ä–æ–∫–∞ (JSON)"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    course_slug = data['course_slug']
    module_id = data['module_id']
    lesson_id = data['lesson_id']
    
    from data import get_lesson_by_id, update_lesson
    
    lesson = get_lesson_by_id(course_slug, module_id, lesson_id)
    if lesson:
        text_content = message.text.strip()
        lesson['text_content'] = "" if text_content == "-" else text_content
        update_lesson(course_slug, module_id, lesson_id, lesson)
        
        if text_content == "-":
            msg = "‚úÖ –¢–µ–∫—Å—Ç–æ–≤–æ–µ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏–µ —É–¥–∞–ª–µ–Ω–æ –∏–∑ —É—Ä–æ–∫–∞!"
        else:
            msg = "‚úÖ –¢–µ–∫—Å—Ç–æ–≤–æ–µ —Å–æ–¥–µ—Ä–∂–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–æ!"
        
        await message.answer(
            msg,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="‚óÄÔ∏è –ö —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{course_slug}_{module_id}_{lesson_id}")]
            ])
        )
    else:
        await message.answer("‚ùå –£—Ä–æ–∫ –Ω–µ –Ω–∞–π–¥–µ–Ω")
    
    await state.clear()


# ===== –°–¢–ê–†–´–ï –û–ë–†–ê–ë–û–¢–ß–ò–ö–ò –î–õ–Ø –ë–î (DEPRECATED) =====

@router.callback_query(F.data.startswith("edit_title_"))
async def edit_lesson_title_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –Ω–∞–∑–≤–∞–Ω–∏—è (DEPRECATED - —Ç–æ–ª—å–∫–æ –¥–ª—è –ë–î)"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —ç—Ç–æ —Å—Ç–∞—Ä—ã–π —Ñ–æ—Ä–º–∞—Ç (—Ç–æ–ª—å–∫–æ —á–∏—Å–ª–æ) –∏–ª–∏ –Ω–æ–≤—ã–π
    try:
        lesson_id = int(callback.data.split("_")[2])
        await callback.answer("‚ö†Ô∏è –ò—Å–ø–æ–ª—å–∑—É–π—Ç–µ –Ω–æ–≤—ã–π –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å —á–µ—Ä–µ–∑ '–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫—É—Ä—Å–∞–º–∏'", show_alert=True)
        return
    except:
        pass
    
    await state.update_data(editing_lesson_id=lesson_id)
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        
        cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await callback.message.edit_text(
            f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –Ω–∞–∑–≤–∞–Ω–∏—è —É—Ä–æ–∫–∞</b>\n\n"
            f"–¢–µ–∫—É—â–µ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ: <b>{lesson.title}</b>\n\n"
            f"–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ:",
            reply_markup=cancel_keyboard
        )
        
        await state.set_state(CourseManagement.editing_lesson_title)
    
    finally:
        db.close()
    
    await callback.answer()


@router.message(CourseManagement.editing_lesson_title)
async def edit_lesson_title_save(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –Ω–∞–∑–≤–∞–Ω–∏—è"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    lesson_id = data['editing_lesson_id']
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        lesson.title = message.text.strip()
        db.commit()
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await message.answer(
            f"‚úÖ –ù–∞–∑–≤–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–æ!\n\n"
            f"–ù–æ–≤–æ–µ –Ω–∞–∑–≤–∞–Ω–∏–µ: <b>{lesson.title}</b>",
            reply_markup=back_keyboard
        )
    
    finally:
        db.close()
        await state.clear()


@router.callback_query(F.data.startswith("edit_desc_"))
async def edit_lesson_desc_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –æ–ø–∏—Å–∞–Ω–∏—è"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    lesson_id = int(callback.data.split("_")[2])
    await state.update_data(editing_lesson_id=lesson_id)
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        
        cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        current_desc = lesson.description if lesson.description else "–ù–µ —É–∫–∞–∑–∞–Ω–æ"
        
        await callback.message.edit_text(
            f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –æ–ø–∏—Å–∞–Ω–∏—è —É—Ä–æ–∫–∞</b>\n\n"
            f"–¢–µ–∫—É—â–µ–µ –æ–ø–∏—Å–∞–Ω–∏–µ:\n{current_desc}\n\n"
            f"–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤–æ–µ –æ–ø–∏—Å–∞–Ω–∏–µ:",
            reply_markup=cancel_keyboard
        )
        
        await state.set_state(CourseManagement.editing_lesson_description)
    
    finally:
        db.close()
    
    await callback.answer()


@router.message(CourseManagement.editing_lesson_description)
async def edit_lesson_desc_save(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –æ–ø–∏—Å–∞–Ω–∏—è"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    lesson_id = data['editing_lesson_id']
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        lesson.description = message.text.strip()
        db.commit()
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await message.answer(
            "‚úÖ –û–ø–∏—Å–∞–Ω–∏–µ —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–æ!",
            reply_markup=back_keyboard
        )
    
    finally:
        db.close()
        await state.clear()


@router.callback_query(F.data.startswith("edit_content_"))
async def edit_lesson_content_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∫–æ–Ω—Ç–µ–Ω—Ç–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    lesson_id = int(callback.data.split("_")[2])
    await state.update_data(editing_lesson_id=lesson_id)
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        
        cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await callback.message.edit_text(
            f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫–æ–Ω—Ç–µ–Ω—Ç–∞ —É—Ä–æ–∫–∞</b>\n\n"
            f"–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤—ã–π —Ç–µ–∫—Å—Ç —É—Ä–æ–∫–∞:",
            reply_markup=cancel_keyboard
        )
        
        await state.set_state(CourseManagement.editing_lesson_content)
    
    finally:
        db.close()
    
    await callback.answer()


@router.message(CourseManagement.editing_lesson_content)
async def edit_lesson_content_save(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –∫–æ–Ω—Ç–µ–Ω—Ç–∞"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    lesson_id = data['editing_lesson_id']
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        lesson.content = message.text.strip()
        db.commit()
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await message.answer(
            "‚úÖ –ö–æ–Ω—Ç–µ–Ω—Ç —É—Ä–æ–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω!",
            reply_markup=back_keyboard
        )
    
    finally:
        db.close()
        await state.clear()


@router.callback_query(F.data.startswith("edit_video_"))
async def edit_lesson_video_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –≤–∏–¥–µ–æ"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    lesson_id = int(callback.data.split("_")[2])
    await state.update_data(editing_lesson_id=lesson_id)
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        
        cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        current_video = lesson.video_url if lesson.video_url else "–ù–µ —É–∫–∞–∑–∞–Ω–æ"
        
        await callback.message.edit_text(
            f"‚úèÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –≤–∏–¥–µ–æ —É—Ä–æ–∫–∞</b>\n\n"
            f"–¢–µ–∫—É—â–µ–µ –≤–∏–¥–µ–æ: {current_video}\n\n"
            f"–û—Ç–ø—Ä–∞–≤—å—Ç–µ –Ω–æ–≤—É—é —Å—Å—ã–ª–∫—É –Ω–∞ –≤–∏–¥–µ–æ:",
            reply_markup=cancel_keyboard
        )
        
        await state.set_state(CourseManagement.editing_lesson_video)
    
    finally:
        db.close()
    
    await callback.answer()


@router.message(CourseManagement.editing_lesson_video)
async def edit_lesson_video_save(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –≤–∏–¥–µ–æ"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    lesson_id = data['editing_lesson_id']
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        lesson.video_url = message.text.strip()
        db.commit()
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ —É—Ä–æ–∫—É", callback_data=f"edit_lesson_{lesson_id}")]
        ])
        
        await message.answer(
            "‚úÖ –°—Å—ã–ª–∫–∞ –Ω–∞ –≤–∏–¥–µ–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∞!",
            reply_markup=back_keyboard
        )
    
    finally:
        db.close()
        await state.clear()


@router.callback_query(F.data.startswith("delete_lesson_"))
async def delete_lesson_confirm(callback: CallbackQuery):
    """–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è —É—Ä–æ–∫–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    lesson_id = int(callback.data.split("_")[2])
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        
        buttons = [
            [InlineKeyboardButton(text="‚úÖ –î–∞, —É–¥–∞–ª–∏—Ç—å", callback_data=f"confirm_delete_{lesson_id}")],
            [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data=f"edit_lesson_{lesson_id}")]
        ]
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        
        await callback.message.edit_text(
            f"‚ö†Ô∏è <b>–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ —É–¥–∞–ª–µ–Ω–∏—è</b>\n\n"
            f"–í—ã –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–æ —Ö–æ—Ç–∏—Ç–µ —É–¥–∞–ª–∏—Ç—å —É—Ä–æ–∫?\n\n"
            f"üìç –ú{lesson.module_number}.–£{lesson.lesson_number}: {lesson.title}\n\n"
            f"–≠—Ç–æ –¥–µ–π—Å—Ç–≤–∏–µ –Ω–µ–ª—å–∑—è –æ—Ç–º–µ–Ω–∏—Ç—å!",
            reply_markup=keyboard
        )
    
    finally:
        db.close()
    
    await callback.answer()


@router.callback_query(F.data.startswith("confirm_delete_"))
async def delete_lesson(callback: CallbackQuery):
    """–£–¥–∞–ª–µ–Ω–∏–µ —É—Ä–æ–∫–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    lesson_id = int(callback.data.split("_")[2])
    
    db: Session = get_db()
    try:
        lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
        course_id = lesson.course_id
        
        db.delete(lesson)
        db.commit()
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ –∫—É—Ä—Å—É", callback_data=f"manage_course_{course_id}")]
        ])
        
        await callback.message.edit_text(
            "‚úÖ <b>–£—Ä–æ–∫ —É—Å–ø–µ—à–Ω–æ —É–¥–∞–ª–µ–Ω!</b>",
            reply_markup=back_keyboard
        )
    
    finally:
        db.close()
    
    await callback.answer()


@router.callback_query(F.data == "download_analytics")
async def download_analytics(callback: CallbackQuery):
    """–ì–µ–Ω–µ—Ä–∞—Ü–∏—è –∏ –æ—Ç–ø—Ä–∞–≤–∫–∞ Excel —Ñ–∞–π–ª–∞ —Å –∞–Ω–∞–ª–∏—Ç–∏–∫–æ–π"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    await callback.answer("‚è≥ –ì–µ–Ω–µ—Ä–∏—Ä—É—é —Ñ–∞–π–ª...")
    
    db: Session = get_db()
    
    try:
        # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º —Ñ—É–Ω–∫—Ü–∏–∏ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å –¥–∞–Ω–Ω—ã–º–∏
        from data import get_all_guides
        
        # –°–æ–∑–¥–∞–µ–º Excel —Ñ–∞–π–ª
        wb = Workbook()
        
        # –õ–∏—Å—Ç 1: –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        ws_stats = wb.active
        ws_stats.title = "–û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞"
        
        # –ó–∞–≥–æ–ª–æ–≤–∫–∏
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws_stats['A1'] = '–û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –±–æ—Ç–∞'
        ws_stats['A1'].font = Font(bold=True, size=14)
        ws_stats.merge_cells('A1:B1')
        
        row = 3
        ws_stats[f'A{row}'] = '–ü–æ–∫–∞–∑–∞—Ç–µ–ª—å'
        ws_stats[f'B{row}'] = '–ó–Ω–∞—á–µ–Ω–∏–µ'
        ws_stats[f'A{row}'].fill = header_fill
        ws_stats[f'B{row}'].fill = header_fill
        ws_stats[f'A{row}'].font = header_font
        ws_stats[f'B{row}'].font = header_font
        
        # –î–∞–Ω–Ω—ã–µ
        total_users = db.query(User).count()
        week_ago = datetime.utcnow() - timedelta(days=7)
        month_ago = datetime.utcnow() - timedelta(days=30)
        
        active_week = db.query(User).filter(User.last_activity >= week_ago).count()
        active_month = db.query(User).filter(User.last_activity >= month_ago).count()
        new_week = db.query(User).filter(User.created_at >= week_ago).count()
        new_month = db.query(User).filter(User.created_at >= month_ago).count()
        
        total_payments = db.query(Payment).filter(Payment.status == 'succeeded').count()
        total_revenue = db.query(func.sum(Payment.amount)).filter(Payment.status == 'succeeded').scalar() or 0
        week_payments = db.query(Payment).filter(
            Payment.status == 'succeeded',
            Payment.created_at >= week_ago
        ).count()
        week_revenue = db.query(func.sum(Payment.amount)).filter(
            Payment.status == 'succeeded',
            Payment.created_at >= week_ago
        ).scalar() or 0
        month_payments = db.query(Payment).filter(
            Payment.status == 'succeeded',
            Payment.created_at >= month_ago
        ).count()
        month_revenue = db.query(func.sum(Payment.amount)).filter(
            Payment.status == 'succeeded',
            Payment.created_at >= month_ago
        ).scalar() or 0
        
        stats_data = [
            ('–ü–û–õ–¨–ó–û–í–ê–¢–ï–õ–ò', ''),
            ('–í—Å–µ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π', total_users),
            ('–ê–∫—Ç–∏–≤–Ω—ã—Ö –∑–∞ –Ω–µ–¥–µ–ª—é', active_week),
            ('–ê–∫—Ç–∏–≤–Ω—ã—Ö –∑–∞ –º–µ—Å—è—Ü', active_month),
            ('–ù–æ–≤—ã—Ö –∑–∞ –Ω–µ–¥–µ–ª—é', new_week),
            ('–ù–æ–≤—ã—Ö –∑–∞ –º–µ—Å—è—Ü', new_month),
            ('', ''),
            ('–§–ò–ù–ê–ù–°–´', ''),
            ('–í—Å–µ–≥–æ –ø–æ–∫—É–ø–æ–∫', total_payments),
            ('–û–±—â–∞—è –≤—ã—Ä—É—á–∫–∞, ‚ÇΩ', f'{total_revenue:,.2f}'),
            ('–ü–æ–∫—É–ø–æ–∫ –∑–∞ –Ω–µ–¥–µ–ª—é', week_payments),
            ('–í—ã—Ä—É—á–∫–∞ –∑–∞ –Ω–µ–¥–µ–ª—é, ‚ÇΩ', f'{week_revenue:,.2f}'),
            ('–ü–æ–∫—É–ø–æ–∫ –∑–∞ –º–µ—Å—è—Ü', month_payments),
            ('–í—ã—Ä—É—á–∫–∞ –∑–∞ –º–µ—Å—è—Ü, ‚ÇΩ', f'{month_revenue:,.2f}'),
            ('', ''),
            ('–ö–û–ù–¢–ï–ù–¢', ''),
            ('–ö—É—Ä—Å–æ–≤', len(get_all_courses())),
            ('–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π', len(get_all_consultations())),
            ('–ì–∞–π–¥–æ–≤', len(get_all_guides())),
        ]
        
        row = 4
        for label, value in stats_data:
            ws_stats[f'A{row}'] = label
            ws_stats[f'B{row}'] = value
            if label and not value:
                ws_stats[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞ –∫–æ–ª–æ–Ω–æ–∫
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        # –õ–∏—Å—Ç 2: –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏
        ws_users = wb.create_sheet("–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏")
        
        headers = ['ID', 'Telegram ID', 'Username', '–ò–º—è', '–§–∞–º–∏–ª–∏—è', '–î–∞—Ç–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏', '–ü–æ—Å–ª–µ–¥–Ω—è—è –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å', '–ü–æ–∫—É–ø–æ–∫']
        for col, header in enumerate(headers, 1):
            cell = ws_users.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        users = db.query(User).order_by(User.created_at.desc()).all()
        
        for row, user in enumerate(users, 2):
            payments_count = db.query(Payment).filter(
                Payment.user_id == user.id,
                Payment.status == 'succeeded'
            ).count()
            
            ws_users.cell(row, 1, user.id)
            ws_users.cell(row, 2, user.telegram_id)
            ws_users.cell(row, 3, user.username or '-')
            ws_users.cell(row, 4, user.first_name or '-')
            ws_users.cell(row, 5, user.last_name or '-')
            ws_users.cell(row, 6, user.created_at.strftime('%d.%m.%Y %H:%M'))
            ws_users.cell(row, 7, user.last_activity.strftime('%d.%m.%Y %H:%M') if user.last_activity else '-')
            ws_users.cell(row, 8, payments_count)
        
        # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws_users.column_dimensions[col].width = 15
        
        # –õ–∏—Å—Ç 3: –ü–æ–∫—É–ø–∫–∏
        ws_payments = wb.create_sheet("–ü–æ–∫—É–ø–∫–∏")
        
        headers = ['ID', '–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å', '–ü—Ä–æ–¥—É–∫—Ç', '–°—É–º–º–∞, ‚ÇΩ', '–°—Ç–∞—Ç—É—Å', '–î–∞—Ç–∞ —Å–æ–∑–¥–∞–Ω–∏—è', '–î–∞—Ç–∞ –æ–ø–ª–∞—Ç—ã']
        for col, header in enumerate(headers, 1):
            cell = ws_payments.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        payments = db.query(Payment).order_by(Payment.created_at.desc()).all()
        
        for row, payment in enumerate(payments, 2):
            user = payment.user
            username = user.username if user.username else f"{user.first_name or '–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å'}"
            
            product_name = '-'
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø –ø—Ä–æ–¥—É–∫—Ç–∞ –∏ –Ω–∞–∑–≤–∞–Ω–∏–µ –∏–∑ –Ω–æ–≤–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä—ã
            from data import get_course_by_slug, get_consultation_by_slug, get_guide_by_id
            
            if payment.product_type == 'course' and payment.course_slug:
                course = get_course_by_slug(payment.course_slug)
                product_name = f"–ö—É—Ä—Å: {course['name']}" if course else "–ö—É—Ä—Å"
            elif payment.product_type == 'consultation' and payment.consultation_slug:
                consultation = get_consultation_by_slug(payment.consultation_slug)
                product_name = f"–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è: {consultation['name']}" if consultation else "–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è"
            elif payment.product_type == 'guide' and payment.product_id:
                guide = get_guide_by_id(payment.product_id)
                product_name = f"–ì–∞–π–¥: {guide['name']}" if guide else "–ì–∞–π–¥"
            else:
                product_name = payment.product_type or "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–æ"
            
            ws_payments.cell(row, 1, payment.id)
            ws_payments.cell(row, 2, username)
            ws_payments.cell(row, 3, product_name)
            ws_payments.cell(row, 4, payment.amount)
            ws_payments.cell(row, 5, payment.status)
            ws_payments.cell(row, 6, payment.created_at.strftime('%d.%m.%Y %H:%M'))
            ws_payments.cell(row, 7, payment.paid_at.strftime('%d.%m.%Y %H:%M') if payment.paid_at else '-')
        
        # –ê–≤—Ç–æ—à–∏—Ä–∏–Ω–∞
        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G'], [8, 20, 30, 12, 12, 18, 18]):
            ws_payments.column_dimensions[col].width = width
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –±—É—Ñ–µ—Ä
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –∏–º—è —Ñ–∞–π–ª–∞ —Å –¥–∞—Ç–æ–π
        filename = f"analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤—Ä–µ–º–µ–Ω–Ω–æ
        temp_path = f"/tmp/{filename}"
        with open(temp_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Ñ–∞–π–ª
        await callback.message.answer_document(
            document=FSInputFile(temp_path, filename=filename),
            caption="üìä <b>–î–µ—Ç–∞–ª—å–Ω–∞—è –∞–Ω–∞–ª–∏—Ç–∏–∫–∞</b>\n\n"
                   "–§–∞–π–ª —Å–æ–¥–µ—Ä–∂–∏—Ç:\n"
                   "‚Ä¢ –û–±—â—É—é —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É\n"
                   "‚Ä¢ –°–ø–∏—Å–æ–∫ –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π\n"
                   "‚Ä¢ –í—Å–µ –ø–æ–∫—É–ø–∫–∏"
        )
        
        # –£–¥–∞–ª—è–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–π —Ñ–∞–π–ª
        os.remove(temp_path)
        
        await callback.answer("‚úÖ –§–∞–π–ª –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω!")
    
    except Exception as e:
        await callback.answer(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏: {str(e)}", show_alert=True)
    
    finally:
        db.close()


@router.callback_query(F.data == "admin_consultations")
async def show_consultations_management(callback: CallbackQuery):
    """–ü–æ–∫–∞–∑–∞—Ç—å —É–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è–º–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    consultations = get_all_consultations()
    
    if not consultations:
        await callback.message.edit_text(
            "üîÆ <b>–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ –æ—Ç—Å—É—Ç—Å—Ç–≤—É—é—Ç</b>\n\n"
            "–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ —Ö—Ä–∞–Ω—è—Ç—Å—è –≤ data/consultations.json",
            reply_markup=get_back_to_admin_keyboard()
        )
    else:
        buttons = []
        for cons in consultations:
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–µ–∫—Å—Ç —Ü–µ–Ω—ã
            if cons.get('options'):
                price_text = "–í–∞—Ä–∏–∞–Ω—Ç—ã"
            elif cons.get('price'):
                price_text = f"{cons['price']:,.0f} ‚ÇΩ"
            else:
                price_text = "–¶–µ–Ω–∞ –Ω–µ —É–∫–∞–∑–∞–Ω–∞"
            
            status_emoji = "‚úÖ" if cons.get('is_active', True) else "‚ùå"
            buttons.append([InlineKeyboardButton(
                text=f"{cons.get('emoji', 'üîÆ')} {cons['name']} ({price_text}) {status_emoji}",
                callback_data=f"manage_consultation_{cons['slug']}"
            )])
        
        buttons.append([InlineKeyboardButton(
            text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å",
            callback_data="admin_panel"
        )])
        
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        
        await callback.message.edit_text(
            "üîÆ <b>–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è–º–∏</b>\n\n"
            "–ü—Ä–æ—Å–º–æ—Ç—Ä –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π. –î–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∏—Å–ø–æ–ª—å–∑—É–π—Ç–µ —Ñ–∞–π–ª data/consultations.json",
            reply_markup=keyboard
        )
    
    await callback.answer()


@router.callback_query(F.data.startswith("manage_consultation_"))
async def manage_consultation(callback: CallbackQuery):
    """–ü—Ä–æ—Å–º–æ—Ç—Ä –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–π –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    from data import get_consultation_by_slug
    
    cons_slug = callback.data.replace("manage_consultation_", "")
    cons = get_consultation_by_slug(cons_slug)
    
    if not cons:
        await callback.answer("–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–∞", show_alert=True)
        return
    
    text = f"üîÆ <b>{cons['name']}</b>\n\n"
    text += f"üìç Slug: <code>{cons['slug']}</code>\n"
    
    if cons.get('price'):
        text += f"üí∞ –¶–µ–Ω–∞: {cons['price']:,.0f} ‚ÇΩ\n"
    
    if cons.get('duration'):
        text += f"‚è± –î–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å: {cons['duration']}\n"
    
    if cons.get('category'):
        text += f"üìÇ –ö–∞—Ç–µ–≥–æ—Ä–∏—è: {cons['category']}\n"
    
    text += f"‚úÖ –ê–∫—Ç–∏–≤–Ω–∞: {'–î–∞' if cons.get('is_active', True) else '–ù–µ—Ç'}\n"
    
    if cons.get('options'):
        text += f"\nüìã –í–∞—Ä–∏–∞–Ω—Ç–æ–≤: {len(cons['options'])}\n"
        for opt in cons['options']:
            if opt.get('is_active', True):
                text += f"  ‚Ä¢ {opt['name']}: {opt['price']:,.0f} ‚ÇΩ\n"
    
    text += "\n\n<i>–î–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è –∏–∑–º–µ–Ω–∏—Ç–µ data/consultations.json</i>"
    
    buttons = [
        [InlineKeyboardButton(text="‚óÄÔ∏è –ù–∞–∑–∞–¥ –∫ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è–º", callback_data="admin_consultations")]
    ]
    
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await callback.message.edit_text(text, reply_markup=keyboard)
    await callback.answer()


@router.callback_query(F.data.startswith("toggle_cons_"))
async def toggle_consultation(callback: CallbackQuery):
    """–ü–µ—Ä–µ–∫–ª—é—á–∏—Ç—å –∞–∫—Ç–∏–≤–Ω–æ—Å—Ç—å –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ - –∑–∞–≥–ª—É—à–∫–∞"""
    await callback.answer("‚ö†Ô∏è –†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ —á–µ—Ä–µ–∑ data/consultations.json", show_alert=True)


# –†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π —á–µ—Ä–µ–∑ JSON —Ñ–∞–π–ª—ã

@router.callback_query(F.data.startswith("edit_cons_"))
async def edit_consultation_help(callback: CallbackQuery):
    """–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π"""
    await callback.message.answer(
        "‚ÑπÔ∏è <b>–†–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π</b>\n\n"
        "–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ —Ö—Ä–∞–Ω—è—Ç—Å—è –≤ —Ñ–∞–π–ª–µ:\n"
        "<code>data/consultations.json</code>\n\n"
        "–î–ª—è —Ä–µ–¥–∞–∫—Ç–∏—Ä–æ–≤–∞–Ω–∏—è:\n"
        "1. –û—Ç–∫—Ä–æ–π—Ç–µ —Ñ–∞–π–ª –≤ —Ä–µ–¥–∞–∫—Ç–æ—Ä–µ\n"
        "2. –ò–∑–º–µ–Ω–∏—Ç–µ –Ω—É–∂–Ω—ã–µ –ø–æ–ª—è\n"
        "3. –°–æ—Ö—Ä–∞–Ω–∏—Ç–µ —Ñ–∞–π–ª\n"
        "4. –ü–µ—Ä–µ–∑–∞–ø—É—Å—Ç–∏—Ç–µ –±–æ—Ç–∞\n\n"
        "üìù –î–æ—Å—Ç—É–ø–Ω—ã–µ –ø–æ–ª—è:\n"
        "‚Ä¢ name - –Ω–∞–∑–≤–∞–Ω–∏–µ\n"
        "‚Ä¢ price - —Ü–µ–Ω–∞\n"
        "‚Ä¢ duration - –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å\n"
        "‚Ä¢ description - –æ–ø–∏—Å–∞–Ω–∏–µ\n"
        "‚Ä¢ is_active - –∞–∫—Ç–∏–≤–Ω–∞/–Ω–µ–∞–∫—Ç–∏–≤–Ω–∞\n"
        "‚Ä¢ options - –≤–∞—Ä–∏–∞–Ω—Ç—ã –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("create_consultation"))
async def create_consultation_help(callback: CallbackQuery):
    """–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –æ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π"""
    await callback.message.edit_text(
        "‚ÑπÔ∏è <b>–°–æ–∑–¥–∞–Ω–∏–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏</b>\n\n"
        "–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏ —Ö—Ä–∞–Ω—è—Ç—Å—è –≤ —Ñ–∞–π–ª–µ:\n"
        "<code>data/consultations.json</code>\n\n"
        "–î–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –Ω–æ–≤–æ–π –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏:\n"
        "1. –û—Ç–∫—Ä–æ–π—Ç–µ —Ñ–∞–π–ª –≤ —Ä–µ–¥–∞–∫—Ç–æ—Ä–µ\n"
        "2. –°–∫–æ–ø–∏—Ä—É–π—Ç–µ —Å—Ç—Ä—É–∫—Ç—É—Ä—É —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–π –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏\n"
        "3. –ò–∑–º–µ–Ω–∏—Ç–µ –ø–æ–ª—è –ø–æ–¥ –Ω–æ–≤—É—é –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—é\n"
        "4. –ù–µ –∑–∞–±—É–¥—å—Ç–µ —É–Ω–∏–∫–∞–ª—å–Ω—ã–π slug\n"
        "5. –°–æ—Ö—Ä–∞–Ω–∏—Ç–µ —Ñ–∞–π–ª\n"
        "6. –ü–µ—Ä–µ–∑–∞–ø—É—Å—Ç–∏—Ç–µ –±–æ—Ç–∞\n\n"
        "üìã –ü—Ä–∏–º–µ—Ä —Å—Ç—Ä—É–∫—Ç—É—Ä—ã –≤ –¥–æ–∫—É–º–µ–Ω—Ç–∞—Ü–∏–∏ DATA_MANAGEMENT.md",
        reply_markup=get_back_to_admin_keyboard()
    )
    await callback.answer()


# –£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –≥–∞–π–¥–∞–º–∏ –≤—ã–Ω–µ—Å–µ–Ω–æ –≤ –æ—Ç–¥–µ–ª—å–Ω—ã–π –º–æ–¥—É–ª—å handlers/admin_guides.py


@router.callback_query(F.data == "create_course")
async def create_course_start(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Å–æ–∑–¥–∞–Ω–∏—è –Ω–æ–≤–æ–≥–æ –∫—É—Ä—Å–∞"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –î–æ—Å—Ç—É–ø –∑–∞–ø—Ä–µ—â–µ–Ω", show_alert=True)
        return
    
    cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data="admin_courses")]
    ])
    
    await callback.message.edit_text(
        "‚ûï <b>–°–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –∫—É—Ä—Å–∞</b>\n\n"
        "–®–∞–≥ 1/4: –í–≤–µ–¥–∏—Ç–µ –Ω–∞–∑–≤–∞–Ω–∏–µ –∫—É—Ä—Å–∞:",
        reply_markup=cancel_keyboard
    )
    
    await state.set_state(CourseCreation.waiting_for_name)
    await callback.answer()


@router.message(CourseCreation.waiting_for_name)
async def create_course_name(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ –Ω–∞–∑–≤–∞–Ω–∏—è –∫—É—Ä—Å–∞"""
    if not is_admin(message.from_user.id):
        return
    
    await state.update_data(course_name=message.text.strip())
    
    cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data="admin_courses")]
    ])
    
    await message.answer(
        f"‚úÖ –ù–∞–∑–≤–∞–Ω–∏–µ: <b>{message.text.strip()}</b>\n\n"
        f"–®–∞–≥ 2/4: –í–≤–µ–¥–∏—Ç–µ slug –∫—É—Ä—Å–∞ (–¥–ª—è URL, –Ω–∞–ø—Ä–∏–º–µ—Ä: astro-basics):",
        reply_markup=cancel_keyboard
    )
    
    await state.set_state(CourseCreation.waiting_for_slug)


@router.message(CourseCreation.waiting_for_slug)
async def create_course_slug(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ slug –∫—É—Ä—Å–∞"""
    if not is_admin(message.from_user.id):
        return
    
    slug = message.text.strip().lower().replace(' ', '-')
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å slug
    db: Session = get_db()
    try:
        existing = db.query(Course).filter(Course.slug == slug).first()
        if existing:
            await message.answer("‚ùå –ö—É—Ä—Å —Å —Ç–∞–∫–∏–º slug —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç. –í–≤–µ–¥–∏—Ç–µ –¥—Ä—É–≥–æ–π slug:")
            return
    finally:
        db.close()
    
    await state.update_data(course_slug=slug)
    
    cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data="admin_courses")]
    ])
    
    await message.answer(
        f"‚úÖ Slug: <code>{slug}</code>\n\n"
        f"–®–∞–≥ 3/4: –í–≤–µ–¥–∏—Ç–µ –æ–ø–∏—Å–∞–Ω–∏–µ –∫—É—Ä—Å–∞:",
        reply_markup=cancel_keyboard
    )
    
    await state.set_state(CourseCreation.waiting_for_description)


@router.message(CourseCreation.waiting_for_description)
async def create_course_description(message: Message, state: FSMContext):
    """–ü–æ–ª—É—á–µ–Ω–∏–µ –æ–ø–∏—Å–∞–Ω–∏—è –∫—É—Ä—Å–∞"""
    if not is_admin(message.from_user.id):
        return
    
    await state.update_data(course_description=message.text.strip())
    
    cancel_keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="‚ùå –û—Ç–º–µ–Ω–∞", callback_data="admin_courses")]
    ])
    
    await message.answer(
        f"‚úÖ –û–ø–∏—Å–∞–Ω–∏–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–æ\n\n"
        f"–®–∞–≥ 4/4: –í–≤–µ–¥–∏—Ç–µ –¥–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å –∫—É—Ä—Å–∞ (–Ω–∞–ø—Ä–∏–º–µ—Ä: 3 –º–µ—Å—è—Ü–∞):",
        reply_markup=cancel_keyboard
    )
    
    await state.set_state(CourseCreation.waiting_for_duration)


@router.message(CourseCreation.waiting_for_duration)
async def create_course_save(message: Message, state: FSMContext):
    """–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –∫—É—Ä—Å–∞"""
    if not is_admin(message.from_user.id):
        return
    
    data = await state.get_data()
    
    db: Session = get_db()
    try:
        # –ü–æ–ª—É—á–∞–µ–º –º–∞–∫—Å–∏–º–∞–ª—å–Ω—ã–π order
        max_order = db.query(func.max(Course.order)).scalar() or 0
        
        new_course = Course(
            name=data['course_name'],
            slug=data['course_slug'],
            description=data['course_description'],
            short_description=data['course_description'][:200],  # –ü–µ—Ä–≤—ã–µ 200 —Å–∏–º–≤–æ–ª–æ–≤
            duration=message.text.strip(),
            is_active=True,
            order=max_order + 1
        )
        
        db.add(new_course)
        db.commit()
        db.refresh(new_course)
        
        back_keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="‚ûï –î–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫", callback_data=f"add_lesson_{new_course.id}")],
            [InlineKeyboardButton(text="‚óÄÔ∏è –ö —Å–ø–∏—Å–∫—É –∫—É—Ä—Å–æ–≤", callback_data="admin_courses")]
        ])
        
        await message.answer(
            f"‚úÖ <b>–ö—É—Ä—Å —É—Å–ø–µ—à–Ω–æ —Å–æ–∑–¥–∞–Ω!</b>\n\n"
            f"üìñ {new_course.name}\n"
            f"üîó Slug: <code>{new_course.slug}</code>\n"
            f"‚è± –î–ª–∏—Ç–µ–ª—å–Ω–æ—Å—Ç—å: {new_course.duration}\n\n"
            f"–¢–µ–ø–µ—Ä—å –≤—ã –º–æ–∂–µ—Ç–µ –¥–æ–±–∞–≤–∏—Ç—å —É—Ä–æ–∫–∏ –≤ –∫—É—Ä—Å.",
            reply_markup=back_keyboard
        )
    
    except Exception as e:
        await message.answer(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫—É—Ä—Å–∞: {str(e)}")
    
    finally:
        db.close()
        await state.clear()


# ==================== –°–û–ó–î–ê–ù–ò–ï –ü–õ–ê–¢–ï–ñ–ù–û–ô –°–°–´–õ–ö–ò ====================

@router.callback_query(F.data == "admin_create_payment_link")
async def admin_create_payment_link(callback: CallbackQuery, state: FSMContext):
    """–ù–∞—á–∞–ª–æ —Å–æ–∑–¥–∞–Ω–∏—è –ø–ª–∞—Ç–µ–∂–Ω–æ–π —Å—Å—ã–ª–∫–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –ù–µ—Ç –¥–æ—Å—Ç—É–ø–∞")
        return
    
    await callback.message.edit_text(
        "üí≥ <b>–°–æ–∑–¥–∞–Ω–∏–µ –ø–ª–∞—Ç–µ–∂–Ω–æ–π —Å—Å—ã–ª–∫–∏</b>\n\n"
        "–í–≤–µ–¥–∏—Ç–µ Telegram username –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (—Å @):\n\n"
        "<i>–ù–∞–ø—Ä–∏–º–µ—Ä: @username –∏–ª–∏ 123456789</i>",
        reply_markup=get_back_to_admin_keyboard()
    )
    
    await state.set_state(PaymentLinkCreation.waiting_for_user_id)
    await callback.answer()


@router.message(StateFilter(PaymentLinkCreation.waiting_for_user_id))
async def process_user_id(message: Message, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤–≤–æ–¥–∞ ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    if not is_admin(message.from_user.id):
        return
    
    user_input = message.text.strip()
    db = await get_db()
    user_repo = UserRepository(db)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —ç—Ç–æ ID –∏–ª–∏ username
    if user_input.startswith('@'):
        username = user_input[1:]  # –£–±–∏—Ä–∞–µ–º @
        user = await user_repo.get_by_username(username)
    else:
        try:
            telegram_id = int(user_input)
            user = await user_repo.get_by_telegram_id(telegram_id)
        except ValueError:
            await message.answer("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç. –í–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ–≤–æ–π ID –∏–ª–∏ username —Å @")
            return
    
    if not user:
        await message.answer(
            "‚ùå –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ –±–∞–∑–µ.\n\n"
            "–£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Ö–æ—Ç—è –±—ã —Ä–∞–∑ –∑–∞–ø—É—Å–∫–∞–ª –±–æ—Ç–∞."
        )
        return
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    await state.update_data(
        user_id=str(user.id),
        telegram_id=user.telegram_id,
        user_name=user.first_name or user.username or f"ID{user.telegram_id}"
    )
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ç–∏–ø –ø—Ä–æ–¥—É–∫—Ç–∞ –∫–∞–∫ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è
    await state.update_data(product_type='consultation')
    
    # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º —Å–ø–∏—Å–æ–∫ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π —Å—Ä–∞–∑—É
    consultations = get_all_consultations()
    
    if not consultations:
        await message.answer(
            "‚ùå –ù–µ—Ç –¥–æ—Å—Ç—É–ø–Ω—ã—Ö –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–π",
            reply_markup=get_back_to_admin_keyboard()
        )
        await state.clear()
        return
    
    buttons = []
    for cons in consultations:
        buttons.append([InlineKeyboardButton(
            text=f"{cons.get('emoji', 'üîÆ')} {cons['name']}",
            callback_data=f"paylink_cons_{cons['slug']}"
        )])
    
    buttons.append([InlineKeyboardButton(text="‚óÄÔ∏è –û—Ç–º–µ–Ω–∞", callback_data="admin_panel")])
    keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
    
    await message.answer(
        f"‚úÖ –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–∞–π–¥–µ–Ω: <b>{user.first_name or user.username or '–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å'}</b>\n"
        f"ID: <code>{user.telegram_id}</code>\n\n"
        "üîÆ –í—ã–±–µ—Ä–∏—Ç–µ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—é:",
        reply_markup=keyboard
    )


@router.callback_query(F.data.startswith("paylink_cons_"))
async def process_consultation_selection(callback: CallbackQuery, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—ã–±–æ—Ä–∞ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –ù–µ—Ç –¥–æ—Å—Ç—É–ø–∞")
        return
    
    from data import get_consultation_by_slug
    
    cons_slug = callback.data.replace("paylink_cons_", "")
    consultation = get_consultation_by_slug(cons_slug)
    
    if not consultation:
        await callback.answer("‚ùå –ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–∞", show_alert=True)
        return
    
    await state.update_data(
        consultation_slug=cons_slug,
        consultation_name=consultation['name']
    )
    
    # –ï—Å–ª–∏ –µ—Å—Ç—å –≤–∞—Ä–∏–∞–Ω—Ç—ã - –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –∏—Ö
    options = consultation.get('options', [])
    
    if options:
        buttons = []
        for option in options:
            if option.get('is_active', True):
                buttons.append([InlineKeyboardButton(
                    text=f"{option['name']} - {option['price']:,.0f} ‚ÇΩ",
                    callback_data=f"paylink_option_{option['id']}"
                )])
        
        buttons.append([InlineKeyboardButton(text="‚óÄÔ∏è –û—Ç–º–µ–Ω–∞", callback_data="admin_panel")])
        keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)
        
        await callback.message.edit_text(
            f"üîÆ {consultation['name']}\n\n"
            "–í—ã–±–µ—Ä–∏—Ç–µ –≤–∞—Ä–∏–∞–Ω—Ç:",
            reply_markup=keyboard
        )
    elif consultation.get('price'):
        # –ï—Å—Ç—å —Ñ–∏–∫—Å–∏—Ä–æ–≤–∞–Ω–Ω–∞—è —Ü–µ–Ω–∞ - –∏—Å–ø–æ–ª—å–∑—É–µ–º –µ—ë –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏
        await state.update_data(amount=consultation['price'])
        
        # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏–µ –∏ —Å—Ä–∞–∑—É —Å–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂
        await callback.message.edit_text(
            f"üîÆ {consultation['name']}\n\n"
            f"üí∞ –°—Ç–æ–∏–º–æ—Å—Ç—å: {consultation['price']:,.0f} ‚ÇΩ\n\n"
            f"‚è≥ –°–æ–∑–¥–∞—é –ø–ª–∞—Ç–µ–∂–Ω—É—é —Å—Å—ã–ª–∫—É...",
            reply_markup=None
        )
        
        # –°–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂
        await create_payment_link(callback.message, state)
    else:
        # –ù–µ—Ç –Ω–∏ –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤, –Ω–∏ —Ü–µ–Ω—ã - –ø—Ä–æ—Å–∏–º –≤–≤–µ—Å—Ç–∏ —Å—É–º–º—É –≤—Ä—É—á–Ω—É—é
        await callback.message.edit_text(
            f"üîÆ {consultation['name']}\n\n"
            "–í–≤–µ–¥–∏—Ç–µ —Å—Ç–æ–∏–º–æ—Å—Ç—å –≤ —Ä—É–±–ª—è—Ö:\n\n"
            "<i>–ù–∞–ø—Ä–∏–º–µ—Ä: 5000</i>",
            reply_markup=get_back_to_admin_keyboard()
        )
        await state.set_state(PaymentLinkCreation.waiting_for_amount)
    
    await callback.answer()


@router.callback_query(F.data.startswith("paylink_option_"))
async def process_option_selection(callback: CallbackQuery, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤—ã–±–æ—Ä–∞ –≤–∞—Ä–∏–∞–Ω—Ç–∞ –∫–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏–∏"""
    if not is_admin(callback.from_user.id):
        await callback.answer("‚ùå –ù–µ—Ç –¥–æ—Å—Ç—É–ø–∞")
        return
    
    from data import get_consultation_by_slug, get_consultation_option
    
    option_id = callback.data.replace("paylink_option_", "")
    
    data = await state.get_data()
    consultation_slug = data.get('consultation_slug')
    consultation = get_consultation_by_slug(consultation_slug)
    
    if not consultation:
        await callback.answer("‚ùå –û—à–∏–±–∫–∞", show_alert=True)
        return
    
    option = get_consultation_option(consultation_slug, option_id)
    
    if not option:
        await callback.answer("‚ùå –í–∞—Ä–∏–∞–Ω—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω", show_alert=True)
        return
    
    await state.update_data(
        consultation_option_id=option_id,
        amount=option['price'],
        option_name=option['name']
    )
    
    # –°–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂
    await create_payment_link(callback.message, state)
    await callback.answer()


@router.message(StateFilter(PaymentLinkCreation.waiting_for_amount))
async def process_amount(message: Message, state: FSMContext):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –≤–≤–æ–¥–∞ —Å—É–º–º—ã"""
    if not is_admin(message.from_user.id):
        return
    
    try:
        amount = float(message.text.strip().replace(',', '.').replace(' ', ''))
        
        if amount <= 0:
            await message.answer("‚ùå –°—É–º–º–∞ –¥–æ–ª–∂–Ω–∞ –±—ã—Ç—å –±–æ–ª—å—à–µ 0")
            return
        
        await state.update_data(amount=amount)
        
        # –°–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂
        await create_payment_link(message, state)
    
    except ValueError:
        await message.answer("‚ùå –ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç —Å—É–º–º—ã. –í–≤–µ–¥–∏—Ç–µ —á–∏—Å–ª–æ (–Ω–∞–ø—Ä–∏–º–µ—Ä: 5000)")


async def create_payment_link(message: Message, state: FSMContext):
    """–°–æ–∑–¥–∞–Ω–∏–µ –ø–ª–∞—Ç–µ–∂–Ω–æ–π —Å—Å—ã–ª–∫–∏ –∏ –ø–ª–∞—Ç–µ–∂–∞ –≤ –ë–î"""
    from payments import YooKassaPayment
    
    data = await state.get_data()
    
    user_id_str = data.get('user_id')
    telegram_id = data.get('telegram_id')
    user_name = data.get('user_name')
    product_type = data.get('product_type')
    amount = data.get('amount')
    
    db = await get_db()
    payment_repo = PaymentRepository(db)
    
    try:
        # –§–æ—Ä–º–∏—Ä—É–µ–º –æ–ø–∏—Å–∞–Ω–∏–µ
        if product_type == 'consultation':
            description = f"–ö–æ–Ω—Å—É–ª—å—Ç–∞—Ü–∏—è: {data.get('consultation_name')}"
            if data.get('option_name'):
                description += f" - {data.get('option_name')}"
        elif product_type == 'course':
            description = f"–ö—É—Ä—Å: {data.get('course_name')}"
            if data.get('tariff_name'):
                description += f" - {data.get('tariff_name')}"
        elif product_type == 'guide':
            description = f"–ì–∞–π–¥: {data.get('guide_name', '–ì–∞–π–¥')}"
        else:
            description = "–ü—Ä–æ–¥—É–∫—Ç"
        
        # –°–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂ –≤ –ë–î
        payment = Payment(
            user_id=ObjectId(user_id_str),
            amount=amount,
            status='pending',
            product_type=product_type,
            is_payment_link=True  # –ü–æ–º–µ—á–∞–µ–º –∫–∞–∫ –ø–ª–∞—Ç–µ–∂ –ø–æ —Å—Å—ã–ª–∫–µ
        )
        
        if product_type == 'consultation':
            payment.consultation_slug = data.get('consultation_slug')
            payment.consultation_option_id = data.get('consultation_option_id')
        elif product_type == 'course':
            payment.course_slug = data.get('course_slug')
            payment.tariff_id = data.get('tariff_id')
        elif product_type == 'guide':
            payment.product_id = data.get('guide_id', 'guide-custom')
        
        payment = await payment_repo.create(payment)
        
        # –°–æ–∑–¥–∞–µ–º –ø–ª–∞—Ç–µ–∂ –≤ YooKassa
        yookassa = YooKassaPayment()
        
        bot_info = await message.bot.get_me()
        return_url = f"https://t.me/{bot_info.username}" if bot_info.username else "https://t.me"
        
        payment_result = yookassa.create_payment(
            amount=amount,
            description=description,
            return_url=return_url,
            metadata={
                'payment_db_id': str(payment.id),
                'user_telegram_id': telegram_id
            }
        )
        
        if not payment_result:
            await payment_repo.update(payment.id, {"status": "failed"})
            await message.answer(
                "‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –ø–ª–∞—Ç–µ–∂–∞ –≤ YooKassa",
                reply_markup=get_back_to_admin_keyboard()
            )
            await state.clear()
            return
        
        # –û–±–Ω–æ–≤–ª—è–µ–º –ø–ª–∞—Ç–µ–∂
        await payment_repo.update(payment.id, {
            "payment_id": payment_result['id'],
            "confirmation_url": payment_result['confirmation_url']
        })
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Å—Å—ã–ª–∫—É –∞–¥–º–∏–Ω—É
        await message.answer(
            f"‚úÖ <b>–ü–ª–∞—Ç–µ–∂–Ω–∞—è —Å—Å—ã–ª–∫–∞ —Å–æ–∑–¥–∞–Ω–∞!</b>\n\n"
            f"üë§ –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å: {user_name}\n"
            f"üí∞ –°—É–º–º–∞: {amount:,.0f} ‚ÇΩ\n"
            f"üì¶ –ü—Ä–æ–¥—É–∫—Ç: {description}\n\n"
            f"üîó –°—Å—ã–ª–∫–∞ –Ω–∞ –æ–ø–ª–∞—Ç—É:\n"
            f"<code>{payment_result['confirmation_url']}</code>\n\n"
            f"<i>–û—Ç–ø—Ä–∞–≤—å—Ç–µ —ç—Ç—É —Å—Å—ã–ª–∫—É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é –¥–ª—è –æ–ø–ª–∞—Ç—ã</i>",
            reply_markup=get_back_to_admin_keyboard()
        )
        
        await state.clear()
    
    except Exception as e:
        await message.answer(
            f"‚ùå –û—à–∏–±–∫–∞: {str(e)}",
            reply_markup=get_back_to_admin_keyboard()
        )
        await state.clear()

